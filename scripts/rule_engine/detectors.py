"""Detector registry — type → callable."""
from __future__ import annotations

import importlib
import json as _json
import re
from typing import Callable

import openpyxl

from ._types import Finding, Rule, Severity
from .context import FileContext, ProjectContext


REGISTRY: dict[str, Callable] = {}


def register(name: str):
    def decorator(fn):
        REGISTRY[name] = fn
        return fn
    return decorator


def _line_for(content: str, offset: int) -> int:
    return content[:offset].count("\n") + 1


@register("regex")
def detect_regex(rule: Rule, fc: FileContext, pc: ProjectContext | None) -> list[Finding]:
    pattern = rule.detect.get("pattern")
    if not pattern:
        return []
    flags_str = rule.detect.get("flags", "")
    flags = 0
    if "i" in flags_str.lower():
        flags |= re.IGNORECASE
    if "m" in flags_str.lower():
        flags |= re.MULTILINE
    if "s" in flags_str.lower():
        flags |= re.DOTALL

    findings: list[Finding] = []
    for m in re.finditer(pattern, fc.active_content, flags):
        findings.append(Finding(
            rule_id=rule.id, severity=rule.severity, category=rule.category,
            file=str(fc.path), line=_line_for(fc.active_content, m.start()),
            message=rule.title,
            fix_mechanical=(rule.fix or {}).get("mechanical"),
            fix_prose=(rule.fix or {}).get("prose"),
        ))
    return findings


@register("regex_with_context")
def detect_regex_with_context(rule: Rule, fc: FileContext, pc) -> list[Finding]:
    params = rule.detect.get("params", {})
    pattern = params.get("pattern")
    if not pattern:
        return []
    safe_prefix = params.get("safe_prefix", []) or []
    safe_suffix = params.get("safe_suffix", []) or []
    prefix_window = params.get("prefix_window", 30)
    suffix_window = params.get("suffix_window", 20)

    findings: list[Finding] = []
    content = fc.active_content
    for m in re.finditer(pattern, content):
        prefix = content[max(0, m.start() - prefix_window): m.start()]
        suffix = content[m.end(): m.end() + suffix_window]
        if any(re.search(p, prefix) for p in safe_prefix):
            continue
        if any(re.search(s, suffix) for s in safe_suffix):
            continue
        findings.append(Finding(
            rule_id=rule.id, severity=rule.severity, category=rule.category,
            file=str(fc.path), line=_line_for(content, m.start()),
            message=rule.title,
            fix_mechanical=(rule.fix or {}).get("mechanical"),
            fix_prose=(rule.fix or {}).get("prose"),
        ))
    return findings


@register("regex_pair")
def detect_regex_pair(rule: Rule, fc: FileContext, pc) -> list[Finding]:
    params = rule.detect.get("params", {})
    must_have = params.get("must_have", []) or []
    must_not_have = params.get("must_not_have", []) or []

    content = fc.active_content
    if not all(re.search(p, content) for p in must_have):
        return []
    if any(re.search(p, content) for p in must_not_have):
        return []

    return [Finding(
        rule_id=rule.id, severity=rule.severity, category=rule.category,
        file=str(fc.path), line=1,
        message=rule.title,
        fix_mechanical=(rule.fix or {}).get("mechanical"),
        fix_prose=(rule.fix or {}).get("prose"),
    )]


@register("securestring_argument_outside_chain")
def detect_securestring_argument_outside_chain(
    rule: Rule, fc: FileContext, pc
) -> list[Finding]:
    params = rule.detect.get("params", {})
    chain_args = params.get("chain_marker_args", []) or []
    chain_types = params.get("chain_marker_types", []) or []

    content = fc.active_content

    has_chain = any(
        re.search(rf'Name="{re.escape(arg)}"', content) for arg in chain_args
    )
    if has_chain:
        return []

    # Cross-file chain check: callee NAO declara in_Credenciais mas ALGUM caller
    # passa SecureString derivada de in_Credenciais(...).Item2 ou
    # in_Credenciais("X").Item2 — caso da cadeia Sicoob (REFramework).
    # Se algum caller passa via chain pattern, callee EH parte da chain.
    if pc is not None and pc.root is not None:
        try:
            callee_basename = fc.path.name
            for other in pc.root.rglob("*.xaml"):
                if other == fc.path:
                    continue
                try:
                    other_text = other.read_text(encoding="utf-8-sig")
                except OSError:
                    continue
                if callee_basename not in other_text:
                    continue
                # Look for InArgument SecureString in InvokeWorkflowFile to this callee
                # whose VALUE expression references one of the chain_args
                invoke_re = re.compile(
                    rf'<ui:InvokeWorkflowFile\b[^>]*?\bWorkflowFileName="[^"]*{re.escape(callee_basename)}"[^>]*?>'
                    r'(.*?)'
                    r'</ui:InvokeWorkflowFile>',
                    re.DOTALL,
                )
                for im in invoke_re.finditer(other_text):
                    body = im.group(1)
                    # Heuristic: any InArgument that references a chain_arg
                    for ca in chain_args:
                        if re.search(rf'\b{re.escape(ca)}\s*\(', body):
                            return []  # caller uses chain pattern — callee accepted
        except Exception:
            # Cross-file analysis is best-effort; never block detection
            pass

    findings: list[Finding] = []
    pattern = re.compile(
        r'<x:Property\s+[^>]*Name="([^"]*)"[^>]*Type="((?:In|Out|InOut)Argument\([^)]*SecureString[^)]*\))"',
        re.DOTALL,
    )
    pattern_alt = re.compile(
        r'<x:Property\s+[^>]*Type="((?:In|Out|InOut)Argument\([^)]*SecureString[^)]*\))"[^>]*Name="([^"]*)"',
        re.DOTALL,
    )

    matches = []
    for m in pattern.finditer(content):
        matches.append((m.group(1), m.group(2), m.start()))
    for m in pattern_alt.finditer(content):
        matches.append((m.group(2), m.group(1), m.start()))

    for name, type_decl, offset in matches:
        if any(t in type_decl for t in chain_types):
            continue
        findings.append(Finding(
            rule_id=rule.id, severity=rule.severity, category=rule.category,
            file=str(fc.path), line=_line_for(content, offset),
            message=f"{rule.title}: arg '{name}' tipo {type_decl}",
            fix_mechanical=(rule.fix or {}).get("mechanical"),
            fix_prose=(rule.fix or {}).get("prose"),
        ))
    return findings


@register("duplicate_id")
def detect_duplicate_id(rule: Rule, fc: FileContext, pc) -> list[Finding]:
    from collections import Counter
    params = rule.detect.get("params", {})
    attr = params.get("attribute")
    if not attr:
        return []
    # Skip patterns: IdRefs auto-gerados por Studio (ex: `VisualBasicValue`1_3`)
    # têm scope-local effect — Studio resolve dentro do element, dups
    # tolerados. Layer 2 (Studio Analyzer) silent.
    skip_patterns = params.get("skip_patterns") or []
    skip_re = [re.compile(p) for p in skip_patterns]
    # Negative lookbehind impede match em attrs que TERMINAM com o nome
    # (ex: `MockedActivityIdRef=` é REFERENCE p/ um IdRef, não declaração;
    # só `sap2010:WorkflowViewState.IdRef=` é declaração real).
    pattern = re.compile(rf'(?<![A-Za-z0-9_]){re.escape(attr)}="([^"]+)"')
    values = pattern.findall(fc.active_content)
    counts = Counter(values)
    findings: list[Finding] = []
    for val, n in counts.items():
        if n <= 1:
            continue
        if any(sp.match(val) for sp in skip_re):
            continue
        m = re.search(rf'(?<![A-Za-z0-9_]){re.escape(attr)}="{re.escape(val)}"', fc.active_content)
        line = _line_for(fc.active_content, m.start()) if m else 1
        findings.append(Finding(
            rule_id=rule.id, severity=rule.severity, category=rule.category,
            file=str(fc.path), line=line,
            message=f"{rule.title}: '{val}' aparece {n}x",
            fix_mechanical=(rule.fix or {}).get("mechanical"),
            fix_prose=(rule.fix or {}).get("prose"),
        ))
    return findings


@register("cross_file_args")
def detect_cross_file_args(rule: Rule, fc: FileContext, pc) -> list[Finding]:
    if pc is None:
        return []
    params = rule.detect.get("params", {})
    direction = params.get("direction")
    if direction not in ("caller_extra", "callee_missing"):
        return []
    if "InvokeWorkflowFile" not in fc.active_content:
        return []

    findings: list[Finding] = []
    invoke_pattern = re.compile(
        r'<ui:InvokeWorkflowFile[^>]*WorkflowFileName="([^"]*)"[^>]*>(.*?)</ui:InvokeWorkflowFile>',
        re.DOTALL,
    )
    args_pattern = re.compile(
        r'<ui:InvokeWorkflowFile\.Arguments>(.*?)</ui:InvokeWorkflowFile\.Arguments>',
        re.DOTALL,
    )
    key_pattern = re.compile(r'x:Key="([^"]+)"')
    prop_pattern = re.compile(r'<x:Property[^>]*Name="([^"]+)"')

    content = fc.active_content
    for invoke_match in invoke_pattern.finditer(content):
        target_file = invoke_match.group(1)
        invoke_block = invoke_match.group(0)
        if target_file.startswith("["):
            continue

        normalized = target_file.replace("\\", "/").lstrip("./")
        target_path = pc.root / normalized
        if not target_path.exists():
            continue
        try:
            target_content = target_path.read_text(encoding="utf-8-sig")
        except Exception:
            continue

        declared = set(prop_pattern.findall(target_content))
        if not declared:
            continue

        args_match = args_pattern.search(invoke_block)
        caller_keys: set[str] = set()
        if args_match:
            caller_keys = set(key_pattern.findall(args_match.group(1)))

        line = _line_for(content, invoke_match.start())

        if direction == "caller_extra":
            extra = caller_keys - declared
            if extra:
                findings.append(Finding(
                    rule_id=rule.id, severity=rule.severity, category=rule.category,
                    file=str(fc.path), line=line,
                    message=f"{rule.title}: {target_file} keys extras {sorted(extra)}",
                    fix_mechanical=(rule.fix or {}).get("mechanical"),
                    fix_prose=(rule.fix or {}).get("prose"),
                ))
        elif direction == "callee_missing":
            missing = declared - caller_keys
            if missing:
                findings.append(Finding(
                    rule_id=rule.id, severity=rule.severity, category=rule.category,
                    file=str(fc.path), line=line,
                    message=f"{rule.title}: {target_file} args faltando {sorted(missing)}",
                    fix_mechanical=(rule.fix or {}).get("mechanical"),
                    fix_prose=(rule.fix or {}).get("prose"),
                ))
    return findings


@register("agent_only")
def detect_agent_only(rule: Rule, fc: FileContext, pc) -> list[Finding]:
    """No-op detector — rule is for agent behavior only."""
    return []


def _parse_version(v: str) -> tuple[int, ...]:
    """Parse '23.10.13.0' or '21.10.8' → tuple of ints."""
    return tuple(int(p) for p in re.split(r"[.\-]", v) if p.isdigit())


@register("json_field_check")
def detect_json_field_check(rule: Rule, fc: FileContext, pc) -> list[Finding]:
    if not str(fc.path).endswith(".json"):
        return []
    try:
        data = _json.loads(fc.content)
    except Exception:
        return []

    params = rule.detect.get("params", {})
    field_path = params.get("path", "")
    expected = params.get("expected")
    expected_in = params.get("expected_in")

    parts = field_path.split(".")
    current = data
    for p in parts:
        if not isinstance(current, dict) or p not in current:
            current = None
            break
        current = current[p]

    if current is None and expected is None and expected_in is None:
        return []

    triggered = False
    if expected is not None and current != expected:
        triggered = True
    if expected_in is not None and current not in expected_in:
        triggered = True

    if not triggered:
        return []

    return [Finding(
        rule_id=rule.id, severity=rule.severity, category=rule.category,
        file=str(fc.path), line=1,
        message=f"{rule.title}: {field_path}={current!r}",
        fix_mechanical=(rule.fix or {}).get("mechanical"),
        fix_prose=(rule.fix or {}).get("prose"),
    )]


@register("json_version_compare")
def detect_json_version_compare(rule: Rule, fc: FileContext, pc) -> list[Finding]:
    if not str(fc.path).endswith(".json"):
        return []
    try:
        data = _json.loads(fc.content)
    except Exception:
        return []

    params = rule.detect.get("params", {})
    field_path = params.get("path", "")
    min_v = params.get("min")
    max_v = params.get("max")

    parts = field_path.split(".")
    current = data
    for p in parts:
        if not isinstance(current, dict) or p not in current:
            return []
        current = current[p]
    if not isinstance(current, str):
        return []

    actual = _parse_version(current)
    triggered = False
    if min_v and actual < _parse_version(min_v):
        triggered = True
    if max_v and actual > _parse_version(max_v):
        triggered = True

    if not triggered:
        return []
    return [Finding(
        rule_id=rule.id, severity=rule.severity, category=rule.category,
        file=str(fc.path), line=1,
        message=f"{rule.title}: {field_path}={current}",
        fix_mechanical=(rule.fix or {}).get("mechanical"),
        fix_prose=(rule.fix or {}).get("prose"),
    )]


@register("nuget_version_check")
def detect_nuget_version_check(rule: Rule, fc: FileContext, pc) -> list[Finding]:
    """Pin/min check em project.json::dependencies.

    Param `exact:` (preferred) — exige versão idêntica; qualquer drift (maior
    OU menor) é violação. Usado para pin strict Sicoob (Activity Migrator GA
    ignora pin do projeto e pega latest stable; sem `exact`, version drift
    >pin passava silencioso).

    Param `min:` (legacy/backward compat) — só flagueia se actual < min.

    Quando ambos são declarados, `exact` tem precedência.
    """
    if pc is None:
        return []
    deps = pc.project_json.get("dependencies", {})
    params = rule.detect.get("params", {})
    package = params.get("package")
    min_v = params.get("min")
    exact_v = params.get("exact")
    if not package or package not in deps:
        return []

    raw = deps[package]
    m = re.match(r"\[?([\d.]+)", raw)
    if not m:
        return []
    actual_str = m.group(1)
    actual = _parse_version(actual_str)

    # exact takes precedence over min
    if exact_v:
        if actual == _parse_version(exact_v):
            return []
        violation = f"esperado [{exact_v}], atual {raw}"
    elif min_v:
        if actual >= _parse_version(min_v):
            return []
        violation = f"mínimo {min_v}, atual {raw}"
    else:
        return []

    return [Finding(
        rule_id=rule.id, severity=rule.severity, category=rule.category,
        file=str(pc.root / "project.json"), line=1,
        message=f"{rule.title}: {package}={raw} ({violation})",
        fix_mechanical=(rule.fix or {}).get("mechanical"),
        fix_prose=(rule.fix or {}).get("prose"),
    )]


def _find_config_xlsx(pc):
    """Return first Config.xlsx via fallback chain or None.

    Order:
      1. Inside project: assets/configs/Config_<Role>.xlsx (canonical).
      2. Inside project: assets/Config_<Role>.xlsx (Sicoob shared variant).
      3. Inside project: Data/Config_<Role>.xlsx (deprecated path).
      4. Inside project: Config*.xlsx (root fallback).
      5. Parent dir (multi-package projects): ../assets/configs/ then ../assets/.
         Pick filename matching role hint from project name (Dispatcher/Performer)
         when multiple present.
    """
    if pc is None:
        return None

    inside = (
        list(pc.root.glob("assets/configs/Config_*.xlsx"))
        or list(pc.root.glob("assets/Config_*.xlsx"))
        or list(pc.root.glob("Data/Config_*.xlsx"))
        or list(pc.root.glob("Config*.xlsx"))
    )
    if inside:
        return inside[0]

    parent = pc.root.parent
    parent_candidates = (
        list(parent.glob("assets/configs/Config_*.xlsx"))
        + list(parent.glob("assets/Config_*.xlsx"))
    )
    if not parent_candidates:
        return None

    name_lower = pc.root.name.lower()
    role = "performer" if "performer" in name_lower else (
        "dispatcher" if "dispatcher" in name_lower else None
    )
    if role:
        for c in parent_candidates:
            if role in c.name.lower():
                return c
    return parent_candidates[0]


def _is_main_or_pj(fc) -> bool:
    """Check if FileContext targets project.json or Main.xaml (gate for once-per-project detectors)."""
    return str(fc.path).endswith("project.json") or fc.path.name == "Main.xaml"


@register("config_xlsx_keys")
def detect_config_xlsx_keys(rule: Rule, fc: FileContext, pc) -> list[Finding]:
    if pc is None:
        return []
    if not str(fc.path).lower().endswith(".xaml"):
        return []
    params = rule.detect.get("params", {})
    mode = params.get("mode", "missing")

    config_path = _find_config_xlsx(pc)
    if config_path is None:
        return []

    try:
        wb = openpyxl.load_workbook(config_path, read_only=True, data_only=True)
    except Exception:
        return []

    config_keys: set[str] = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(min_row=2, values_only=True)
        for row in rows:
            if row and row[0]:
                config_keys.add(str(row[0]))
    wb.close()

    used_pairs = re.findall(
        r'in_Config\(&quot;([^&"]+)&quot;\)|in_Config\("([^"]+)"\)',
        fc.active_content,
    )
    used_keys = {k for pair in used_pairs for k in pair if k}

    findings: list[Finding] = []
    if mode == "missing":
        missing = used_keys - config_keys
        if missing:
            findings.append(Finding(
                rule_id=rule.id, severity=rule.severity, category=rule.category,
                file=str(fc.path), line=1,
                message=f"{rule.title}: chaves usadas faltam em Config: {sorted(missing)}",
                fix_mechanical=(rule.fix or {}).get("mechanical"),
                fix_prose=(rule.fix or {}).get("prose"),
            ))
    return findings


@register("python")
def detect_python(rule: Rule, fc: FileContext, pc) -> list[Finding]:
    params = rule.detect.get("params", {})
    module_name = params.get("module")
    function_name = params.get("function")
    if not module_name or not function_name:
        return []
    try:
        mod = importlib.import_module(module_name)
        fn = getattr(mod, function_name)
    except Exception as e:
        return [Finding(
            rule_id=rule.id, severity=Severity.WARN, category=rule.category,
            file=str(fc.path), line=0,
            message=f"[INTERNAL] python detector module/function não encontrado: {e}",
        )]
    return fn(rule, fc, pc)
