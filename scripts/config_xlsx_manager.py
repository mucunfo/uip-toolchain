#!/usr/bin/env python3
"""Config.xlsx manager for UiPath REFramework projects.

Add, list, and validate Config.xlsx entries. Works with the four-sheet
structure: Settings, Constants, Assets, Exceptions.

Config files are discovered under the assets/ directory using the
pattern assets/Config_*.xlsx (e.g., assets/Config_ProcessName.xlsx).

Usage:
    python3 scripts/config_xlsx_manager.py add <project> --sheet Settings --key WebApp_CredentialAssetName --value WebApp_Credential --desc "Credential asset name for WebApp login"
    python3 scripts/config_xlsx_manager.py add <project> --sheet Assets --key WebApp_Url --asset WebApp_ApplicationUrl --folder Shared --desc "the target web application login URL"
    python3 scripts/config_xlsx_manager.py add <project> --sheet Exceptions --key BusinessRuleException_InvalidData --value "Invalid data found" --desc "Thrown when input data fails validation"
    python3 scripts/config_xlsx_manager.py list <project>
    python3 scripts/config_xlsx_manager.py validate <project>
"""

import argparse
import os
import sys

# Ensure UTF-8 output on all platforms (Windows cmd defaults to cp1252)
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def find_config(project_dir: str) -> str:
    """Locate Config_*.xlsx in the assets/ subdirectory of a project directory.

    Searches for files matching the pattern assets/Config_*.xlsx.
    If exactly one match is found it is returned. If multiple matches exist,
    the user is notified and the first (alphabetically sorted) is used.
    """
    import glob

    assets_dir = os.path.join(project_dir, "assets")
    pattern = os.path.join(assets_dir, "Config_*.xlsx")
    matches = sorted(glob.glob(pattern))

    if not matches:
        print(
            f"ERROR: No Config_*.xlsx found in {assets_dir}",
            file=sys.stderr,
        )
        print(
            "  Expected pattern: assets/Config_<ProcessName>.xlsx",
            file=sys.stderr,
        )
        sys.exit(1)

    if len(matches) > 1:
        print(
            f"WARNING: Multiple Config files found, using: {os.path.basename(matches[0])}",
            file=sys.stderr,
        )
        for m in matches:
            print(f"  - {m}", file=sys.stderr)

    return matches[0]


def get_sheet_keys(ws) -> dict[str, int]:
    """Get all Name-column keys and their row numbers from a worksheet."""
    keys = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row[0].value and str(row[0].value).strip():
            keys[str(row[0].value).strip()] = row_idx
    return keys


def cmd_add(args):
    """Add or update a key in Config.xlsx."""
    config_path = find_config(args.project)
    wb = openpyxl.load_workbook(config_path)

    sheet_name = args.sheet
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[sheet_name]
    existing = get_sheet_keys(ws)

    if args.key in existing:
        row = existing[args.key]
        if sheet_name == "Assets":
            ws.cell(row=row, column=2, value=args.asset or "")
            ws.cell(row=row, column=3, value=args.folder or "")
            ws.cell(row=row, column=4, value=args.desc or "")
        else:
            ws.cell(row=row, column=2, value=args.value or "")
            ws.cell(row=row, column=3, value=args.desc or "")
        print(f"Updated '{args.key}' in {sheet_name} sheet (row {row})")
    else:
        # Find first empty row (skip blank separator rows, append at end)
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=1, value=args.key)
        if sheet_name == "Assets":
            ws.cell(row=target_row, column=2, value=args.asset or "")
            ws.cell(row=target_row, column=3, value=args.folder or "")
            ws.cell(row=target_row, column=4, value=args.desc or "")
        else:
            ws.cell(row=target_row, column=2, value=args.value or "")
            ws.cell(row=target_row, column=3, value=args.desc or "")
        print(f"Added '{args.key}' to {sheet_name} sheet (row {target_row})")

    wb.save(config_path)


def cmd_list(args):
    """List all keys in Config.xlsx grouped by sheet."""
    config_path = find_config(args.project)
    wb = openpyxl.load_workbook(config_path, data_only=True)

    total = 0
    for sheet_name in ("Settings", "Constants", "Assets", "Exceptions"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        print(f"\n  {sheet_name} sheet:")

        has_keys = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip():
                key = str(row[0]).strip()
                if sheet_name == "Assets":
                    asset = str(row[1] or "").strip()
                    folder = str(row[2] or "").strip()
                    print(f"    {key:<35} Asset: {asset:<30} Folder: {folder}")
                else:
                    value = str(row[1] if row[1] is not None else "").strip()
                    print(f"    {key:<35} = {value}")
                has_keys = True
                total += 1
        if not has_keys:
            print(f"    (empty)")

    print(f"\n  Total: {total} keys")


def cmd_validate(args):
    """Cross-reference Config.xlsx against XAML Config() references."""
    import re

    config_path = find_config(args.project)
    wb = openpyxl.load_workbook(config_path, data_only=True)

    # Collect defined keys
    defined = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip():
                defined[str(row[0]).strip()] = sheet_name

    # Collect XAML Config() references
    xaml_keys: dict[str, list[str]] = {}
    for root, dirs, files in os.walk(args.project):
        dirs[:] = [d for d in dirs if d not in ("assets", ".local", ".objects")]
        for f in files:
            if not f.endswith(".xaml"):
                continue
            fpath = os.path.join(root, f)
            try:
                with open(fpath, "r", encoding="utf-8-sig") as fh:
                    content = fh.read()
            except Exception:
                continue

            for pattern in [
                r'Config\(&quot;([^&]+)&quot;\)',
                r'in_Config\(&quot;([^&]+)&quot;\)',
                r'Config\("([^"]+)"\)',
                r'in_Config\("([^"]+)"\)',
            ]:
                for key in re.findall(pattern, content):
                    xaml_keys.setdefault(key, []).append(f)

    if not xaml_keys:
        print("No Config() references found in XAML files.")
        return

    # Cross-reference
    missing = {k: v for k, v in xaml_keys.items() if k not in defined}
    unused_framework = {
        "MaxRetryNumber", "MaxConsecutiveSystemExceptions", "ExScreenshotsFolderPath",
        "RetryNumberGetTransactionItem", "RetryNumberSetTransactionStatus",
        "ShouldMarkJobAsFaulted", "LogMessage_GetTransactionData",
        "LogMessage_GetTransactionDataError", "LogMessage_Success",
        "LogMessage_BusinessRuleException", "LogMessage_ApplicationException",
        "ExceptionMessage_ConsecutiveErrors", "logF_BusinessProcessName",
    }
    unused = {k: v for k, v in defined.items()
              if k not in xaml_keys and k not in unused_framework}

    print(f"\n  Config.xlsx: {len(defined)} keys defined")
    print(f"  XAML files:  {len(xaml_keys)} keys referenced")

    if missing:
        print(f"\n  ❌ MISSING from Config.xlsx ({len(missing)} keys):")
        for k in sorted(missing):
            files = ", ".join(sorted(set(missing[k])))
            print(f"    {k:<35} (used in: {files})")
    else:
        print(f"\n  ✅ All {len(xaml_keys)} XAML keys found in Config.xlsx")

    if unused:
        print(f"\n  ⚠️  UNUSED in Config.xlsx ({len(unused)} keys):")
        for k in sorted(unused):
            print(f"    {k:<35} ({unused[k]} sheet)")

    matched = len(xaml_keys) - len(missing)
    print(f"\n  Summary: {matched}/{len(xaml_keys)} matched, "
          f"{len(missing)} missing, {len(unused)} unused")

    sys.exit(1 if missing else 0)


def main():
    parser = argparse.ArgumentParser(
        description="Manage Config.xlsx for UiPath REFramework projects"
    )
    sub = parser.add_subparsers(dest="command", required=True)

    # add
    p_add = sub.add_parser("add", help="Add or update a key in Config.xlsx")
    p_add.add_argument("project", help="Project directory")
    p_add.add_argument("--sheet", required=True,
                       choices=["Settings", "Constants", "Assets", "Exceptions"],
                       help="Target sheet")
    p_add.add_argument("--key", required=True, help="Config key name")
    p_add.add_argument("--value", help="Value (Settings/Constants)")
    p_add.add_argument("--asset", help="Orchestrator Asset name (Assets sheet)")
    p_add.add_argument("--folder", help="Orchestrator folder (Assets sheet)")
    p_add.add_argument("--desc", help="Description")

    # list
    p_list = sub.add_parser("list", help="List all Config.xlsx keys")
    p_list.add_argument("project", help="Project directory")

    # validate
    p_val = sub.add_parser("validate", help="Cross-reference Config.xlsx vs XAML")
    p_val.add_argument("project", help="Project directory")

    args = parser.parse_args()

    if args.command == "add":
        if args.sheet == "Assets" and not args.asset:
            parser.error("--asset required for Assets sheet")
        if args.sheet in ("Settings", "Constants", "Exceptions") and args.value is None:
            parser.error("--value required for Settings/Constants/Exceptions sheets")
        cmd_add(args)
    elif args.command == "list":
        cmd_list(args)
    elif args.command == "validate":
        cmd_validate(args)


if __name__ == "__main__":
    main()
