import json as jsonlib
from pathlib import Path
import pytest
from scripts.rule_engine.detectors import (
    detect_json_field_check, detect_json_version_compare, detect_nuget_version_check,
    REGISTRY,
)
from scripts.rule_engine._types import Rule, Severity
from scripts.rule_engine.context import FileContext, ProjectContext


def make_rule(detect_params, sev=Severity.WARN, category="architectural"):
    return Rule(
        id="X-T", severity=sev, category=category, target="all",
        title="t", description="",
        detect=detect_params,
    )


# ---- Task 21: json_field_check ----

def test_json_field_check_expected_value(tmp_path):
    pj = tmp_path / "project.json"
    pj.write_text(jsonlib.dumps({"studioVersion": "21.10.8.0"}))
    fc = FileContext(pj)
    rule = make_rule({
        "type": "json_field_check",
        "params": {"path": "studioVersion", "expected": "23.10.13"},
    })
    findings = detect_json_field_check(rule, fc, None)
    assert len(findings) == 1


def test_json_field_check_matches_no_finding(tmp_path):
    pj = tmp_path / "project.json"
    pj.write_text(jsonlib.dumps({"studioVersion": "23.10.13"}))
    fc = FileContext(pj)
    rule = make_rule({
        "type": "json_field_check",
        "params": {"path": "studioVersion", "expected": "23.10.13"},
    })
    assert detect_json_field_check(rule, fc, None) == []


def test_json_field_check_in_set(tmp_path):
    pj = tmp_path / "project.json"
    pj.write_text(jsonlib.dumps({"targetFramework": "Bogus"}))
    fc = FileContext(pj)
    rule = make_rule({
        "type": "json_field_check",
        "params": {"path": "targetFramework", "expected_in": ["Windows", "Legacy"]},
    }, sev=Severity.ERROR, category="breaking")
    findings = detect_json_field_check(rule, fc, None)
    assert len(findings) == 1


# ---- Task 22: json_version_compare ----

def test_json_version_compare_below_min(tmp_path):
    pj = tmp_path / "project.json"
    pj.write_text(jsonlib.dumps({"studioVersion": "21.10.8.0"}))
    fc = FileContext(pj)
    rule = make_rule({
        "type": "json_version_compare",
        "params": {"path": "studioVersion", "min": "23.10.0"},
    })
    findings = detect_json_version_compare(rule, fc, None)
    assert len(findings) == 1


def test_json_version_compare_above_min_no_finding(tmp_path):
    pj = tmp_path / "project.json"
    pj.write_text(jsonlib.dumps({"studioVersion": "24.0.0"}))
    fc = FileContext(pj)
    rule = make_rule({
        "type": "json_version_compare",
        "params": {"path": "studioVersion", "min": "23.10.0"},
    })
    assert detect_json_version_compare(rule, fc, None) == []


# ---- Task 23: nuget_version_check ----

def test_nuget_below_min(tmp_path):
    proj = tmp_path / "P"
    proj.mkdir()
    pj = proj / "project.json"
    pj.write_text(jsonlib.dumps({
        "targetFramework": "Windows",
        "dependencies": {"UiPath.UIAutomation.Activities": "[20.0.0]"}
    }))
    f = proj / "Foo.xaml"
    f.write_text("<Activity/>")
    fc = FileContext(f)
    pc = ProjectContext.find_root(f)
    rule = make_rule({
        "type": "nuget_version_check",
        "params": {"package": "UiPath.UIAutomation.Activities", "min": "23.10.0"},
    }, sev=Severity.ERROR, category="breaking")
    findings = detect_nuget_version_check(rule, fc, pc)
    assert len(findings) == 1


def test_nuget_above_min_no_finding(tmp_path):
    proj = tmp_path / "P"
    proj.mkdir()
    pj = proj / "project.json"
    pj.write_text(jsonlib.dumps({
        "targetFramework": "Windows",
        "dependencies": {"UiPath.UIAutomation.Activities": "[25.10.8]"}
    }))
    f = proj / "Foo.xaml"
    f.write_text("<Activity/>")
    fc = FileContext(f)
    pc = ProjectContext.find_root(f)
    rule = make_rule({
        "type": "nuget_version_check",
        "params": {"package": "UiPath.UIAutomation.Activities", "min": "23.10.0"},
    })
    assert detect_nuget_version_check(rule, fc, pc) == []


# ---- Registry ----

def test_registry_has_data_detectors():
    for name in ("json_field_check", "json_version_compare", "nuget_version_check"):
        assert name in REGISTRY


import shutil
import openpyxl
from scripts.rule_engine.detectors import (
    detect_config_xlsx_keys, detect_config_xlsx_sheets,
    detect_config_xlsx_columns, detect_config_xlsx_naming,
)


FIX_DIR = Path(__file__).parent / "fixtures"


# ---- Task 24: config_xlsx_keys ----

def test_config_xlsx_keys_missing(tmp_path):
    proj = tmp_path / "P"
    proj.mkdir()
    (proj / "project.json").write_text('{"targetFramework":"Windows"}')
    cfg_dir = proj / "assets" / "configs"
    cfg_dir.mkdir(parents=True)
    shutil.copy(FIX_DIR / "config_sample.xlsx", cfg_dir / "Config_Performer.xlsx")

    f = proj / "Foo.xaml"
    f.write_text('<Activity>[in_Config(&quot;NonExistentKey&quot;).ToString]</Activity>')
    fc = FileContext(f)
    pc = ProjectContext.find_root(f)

    rule = make_rule({"type": "config_xlsx_keys", "params": {"mode": "missing"}},
                     sev=Severity.ERROR, category="architectural")
    findings = detect_config_xlsx_keys(rule, fc, pc)
    assert len(findings) == 1
    assert "NonExistentKey" in findings[0].message


def test_config_xlsx_keys_existing_no_finding(tmp_path):
    proj = tmp_path / "P"
    proj.mkdir()
    (proj / "project.json").write_text('{"targetFramework":"Windows"}')
    cfg_dir = proj / "assets" / "configs"
    cfg_dir.mkdir(parents=True)
    shutil.copy(FIX_DIR / "config_sample.xlsx", cfg_dir / "Config_Performer.xlsx")

    f = proj / "Foo.xaml"
    f.write_text('<Activity>[in_Config(&quot;KeyA&quot;).ToString]</Activity>')
    fc = FileContext(f)
    pc = ProjectContext.find_root(f)

    rule = make_rule({"type": "config_xlsx_keys", "params": {"mode": "missing"}})
    assert detect_config_xlsx_keys(rule, fc, pc) == []


# ---- Task 25: config_xlsx_sheets / columns / naming ----

def test_config_xlsx_sheets_missing(tmp_path):
    proj = tmp_path / "P"
    proj.mkdir()
    (proj / "project.json").write_text('{"targetFramework":"Windows"}')
    (proj / "assets" / "configs").mkdir(parents=True)
    cfg = proj / "assets" / "configs" / "Config_X.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Settings"
    wb.save(cfg)

    f = proj / "Main.xaml"
    f.write_text("<Activity/>")
    fc = FileContext(f)
    pc = ProjectContext.find_root(f)
    rule = make_rule({"type": "config_xlsx_sheets",
                      "params": {"required": ["Settings", "Constants", "Assets"]}},
                     sev=Severity.WARN, category="architectural")
    findings = detect_config_xlsx_sheets(rule, fc, pc)
    assert any("Constants" in f.message for f in findings)


def test_config_xlsx_sheets_skip_when_not_main(tmp_path):
    proj = tmp_path / "P"
    proj.mkdir()
    (proj / "project.json").write_text('{"targetFramework":"Windows"}')
    f = proj / "Foo.xaml"  # not Main, not project.json
    f.write_text("<Activity/>")
    fc = FileContext(f)
    pc = ProjectContext.find_root(f)
    rule = make_rule({"type": "config_xlsx_sheets",
                      "params": {"required": ["Settings"]}})
    assert detect_config_xlsx_sheets(rule, fc, pc) == []


def test_config_xlsx_columns_missing(tmp_path):
    proj = tmp_path / "P"
    proj.mkdir()
    (proj / "project.json").write_text('{"targetFramework":"Windows"}')
    (proj / "assets" / "configs").mkdir(parents=True)
    cfg = proj / "assets" / "configs" / "Config_X.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Settings"
    ws.append(["Name"])  # Missing Value, Description
    wb.save(cfg)

    f = proj / "Main.xaml"
    f.write_text("<Activity/>")
    fc = FileContext(f)
    pc = ProjectContext.find_root(f)
    rule = make_rule({"type": "config_xlsx_columns",
                      "params": {"sheet": "Settings",
                                 "required_columns": ["Name", "Value", "Description"]}})
    findings = detect_config_xlsx_columns(rule, fc, pc)
    assert len(findings) == 1
    assert "Value" in findings[0].message


def test_config_xlsx_naming_violation(tmp_path):
    proj = tmp_path / "P"
    proj.mkdir()
    (proj / "project.json").write_text('{"targetFramework":"Windows"}')
    (proj / "assets" / "configs").mkdir(parents=True)
    cfg = proj / "assets" / "configs" / "Config_X.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Settings"
    ws.append(["Name", "Value"])
    ws.append(["bad key with spaces", "x"])
    wb.save(cfg)

    f = proj / "Main.xaml"
    f.write_text("<Activity/>")
    fc = FileContext(f)
    pc = ProjectContext.find_root(f)
    rule = make_rule({"type": "config_xlsx_naming",
                      "params": {"pattern": r"^[A-Z][A-Za-z0-9_]+$"}})
    findings = detect_config_xlsx_naming(rule, fc, pc)
    assert len(findings) == 1


def test_registry_has_xlsx_detectors():
    for name in ("config_xlsx_keys", "config_xlsx_sheets",
                 "config_xlsx_columns", "config_xlsx_naming"):
        assert name in REGISTRY
