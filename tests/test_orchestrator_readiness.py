import json
from pathlib import Path

import pytest

from uip_engine import cli
from uip_engine.official_uip import OfficialUipEnvelope, OfficialUipResult
from uip_engine.orchestrator_readiness import (
    audit_orchestrator_readiness,
    format_readiness_result,
    load_readiness_manifest,
)


def _result(data, *, code="Ok", returncode=0, result="Success"):
    payload = {
        "Result": result,
        "Code": code,
        "Data": data,
    }
    envelope = OfficialUipEnvelope(
        result=result,
        code=code,
        data=data,
        message=None,
        instructions=None,
        context=None,
        log=None,
        raw=payload,
    )
    return OfficialUipResult(
        argv=["uip"],
        returncode=returncode,
        stdout=json.dumps(payload),
        stderr="",
        envelope=envelope,
    )


def _write_manifest(path: Path) -> Path:
    path.write_text(
        json.dumps({
            "tenant": "RPA_Desenvolvimento",
            "items": [{
                "packageId": "InvoiceProcessing",
                "packageVersion": "1.2.4",
                "processKey": "11111111-1111-1111-1111-111111111111",
                "folderPath": "Shared",
                "runtimeType": "Unattended",
                "processName": "Invoice Processing",
                "targetFramework": "Windows",
                "entryPointPath": "Main.xaml",
            }],
        }),
        encoding="utf-8",
    )
    return path


def _write_manifest_with_required_inputs(
    path: Path,
    *,
    input_arguments=None,
    input_file: str | None = None,
) -> Path:
    item = {
        "packageId": "InvoiceProcessing",
        "packageVersion": "1.2.4",
        "processKey": "11111111-1111-1111-1111-111111111111",
        "folderPath": "Shared",
        "runtimeType": "Unattended",
        "requiredInputArguments": ["in_RequiredTicket"],
    }
    if input_arguments is not None:
        item["inputArguments"] = input_arguments
    if input_file is not None:
        item["inputFile"] = input_file
    path.write_text(
        json.dumps({
            "tenant": "RPA_Desenvolvimento",
            "items": [item],
        }),
        encoding="utf-8",
    )
    return path


def _write_manifest_with_resources(path: Path) -> Path:
    path.write_text(
        json.dumps({
            "tenant": "RPA_Desenvolvimento",
            "items": [{
                "packageId": "InvoiceProcessing",
                "packageVersion": "1.2.4",
                "processKey": "11111111-1111-1111-1111-111111111111",
                "folderPath": "Shared",
                "runtimeType": "Unattended",
                "requiredResources": [
                    {
                        "type": "asset",
                        "name": "ArquivoConfiguracao_Performer",
                        "valueType": "Text",
                    },
                    {
                        "type": "credential",
                        "name": "CredencialDB2",
                    },
                    {
                        "type": "queue",
                        "name": "Fila de Transacoes",
                    },
                    {
                        "type": "bucket",
                        "name": "anexos",
                    },
                    {
                        "type": "bucketFile",
                        "bucketName": "anexos",
                        "path": "input/config.xlsx",
                    },
                    {
                        "type": "calendar",
                        "name": "Calendario DEV",
                    },
                ],
            }],
        }),
        encoding="utf-8",
    )
    return path


def _write_manifest_with_config_hints(path: Path, config_path: Path) -> Path:
    path.write_text(
        json.dumps({
            "tenant": "RPA_Desenvolvimento",
            "items": [{
                "packageId": "InvoiceProcessing",
                "packageVersion": "1.2.4",
                "processKey": "11111111-1111-1111-1111-111111111111",
                "folderPath": "Shared",
                "runtimeType": "Unattended",
                "requiredResources": [{
                    "type": "asset",
                    "name": "ArquivoConfiguracao_Performer",
                    "valueType": "Text",
                    "source": "source:ArquivoConfiguracao",
                }],
                "runtimeResourceHints": [
                    {"type": "credential", "configKey": "NomeCredencialDB2"},
                    {"type": "queue", "configKey": "OrchestratorQueueName"},
                    {"type": "calendar", "configKey": "CalendarioOrq"},
                ],
            }],
        }),
        encoding="utf-8",
    )
    from openpyxl import Workbook
    workbook = Workbook()
    settings = workbook.active
    settings.title = "Settings"
    settings.append(["Name", "Value", "Description"])
    settings.append(["NomeCredencialDB2", "CredencialDB2_DEV", ""])
    settings.append(["OrchestratorQueueName", "Fila DEV", ""])
    settings.append(["CalendarioOrq", "Calendario DEV", ""])
    assets = workbook.create_sheet("Assets")
    assets.append(["Name", "Asset", "OrchestratorAssetFolder"])
    assets.append(["ApiEndpoint", "ApiEndpoint_DEV", "Shared"])
    credentials = workbook.create_sheet("Credentials")
    credentials.append(["Name", "CredentialAsset", "OrchestratorAssetFolder"])
    credentials.append(["Office365", "CredencialOffice_DEV", "Shared"])
    workbook.save(config_path)
    return path


def _fake_run_ok(calls):
    def fake_run(command):
        calls.append(command)
        if command[:2] == ["login", "status"]:
            return _result({"Status": "Logged in"})
        if command[:3] == ["login", "tenant", "list"]:
            return _result([{"TenantName": "RPA_Desenvolvimento"}])
        if command[:3] == ["login", "tenant", "set"]:
            return _result({"TenantName": command[3]})
        if command[:3] == ["or", "packages", "versions"]:
            return _result([{"Version": "1.2.4"}])
        if command[:3] == ["or", "packages", "entry-points"]:
            return _result([{
                "Path": "Main.xaml",
                "DisplayName": "Main",
                "InputArguments": {
                    "in_RequiredTicket": "String",
                    "invoicePath": "String",
                },
            }])
        if command[:3] == ["or", "processes", "get"]:
            return _result({
                "Key": command[3],
                "Name": "Invoice Processing",
                "ProcessKey": "InvoiceProcessing",
                "ProcessVersion": "1.2.4",
                "FolderPath": "Shared",
                "ProcessType": "Process",
                "Enabled": True,
                "TargetFramework": "Windows",
                "EntryPointPath": "Main.xaml",
            })
        if command[:3] == ["or", "processes", "resources"]:
            return _result([
                {
                    "ResourceType": "Queue",
                    "ResourceName": "Invoices",
                    "ValidationStatus": "Success",
                }
            ])
        if command[:3] == ["or", "folders", "runtimes"]:
            return _result([
                {"Type": "Unattended", "Total": 2, "Connected": 1, "Available": 1}
            ])
        if command[:3] == ["or", "machines", "list"]:
            return _result([
                {
                    "Name": "dev-bot-01",
                    "Key": "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
                    "UnattendedSlots": 1,
                }
            ])
        if command[:4] == ["or", "sessions", "unattended", "list"]:
            return _result([
                {
                    "SessionId": 701,
                    "MachineName": "dev-bot-01",
                    "RuntimeType": "Unattended",
                    "State": "Available",
                    "IsUnresponsive": False,
                }
            ])
        raise AssertionError(command)

    return fake_run


def test_audit_orchestrator_readiness_accepts_valid_manifest(tmp_path):
    manifest = _write_manifest(tmp_path / "manifest.json")
    calls = []

    result = audit_orchestrator_readiness(manifest, run_uip=_fake_run_ok(calls))

    assert result.ok
    assert result.error_count == 0
    assert calls[:3] == [
        ["login", "status", "--output", "json"],
        ["login", "tenant", "list", "--output", "json"],
        ["login", "tenant", "set", "RPA_Desenvolvimento", "--output", "json"],
    ]
    assert [
        "or", "packages", "entry-points",
        "InvoiceProcessing:1.2.4",
        "--output", "json",
    ] in calls
    assert ["or", "processes", "resources", "11111111-1111-1111-1111-111111111111", "--all-fields", "--output", "json"] in calls
    assert ["or", "folders", "runtimes", "Shared", "--output", "json"] in calls
    assert ["or", "machines", "list", "--all-fields", "--limit", "1000", "--output", "json", "--folder-path", "Shared"] in calls
    assert ["or", "sessions", "unattended", "list", "--runtime-type", "Unattended", "--limit", "1000", "--output", "json", "--folder-path", "Shared"] in calls


def test_audit_orchestrator_readiness_rejects_target_framework_mismatch(tmp_path):
    manifest = _write_manifest(tmp_path / "manifest.json")

    def fake_run(command):
        if command[:3] == ["or", "processes", "get"]:
            return _result({
                "Key": command[3],
                "Name": "Invoice Processing",
                "ProcessKey": "InvoiceProcessing",
                "ProcessVersion": "1.2.4",
                "FolderPath": "Shared",
                "TargetFramework": "WindowsLegacy",
                "EntryPointPath": "Main.xaml",
            })
        return _fake_run_ok([])(command)

    result = audit_orchestrator_readiness(manifest, run_uip=fake_run)

    assert not result.ok
    assert any(issue.code == "ORCH-PROCESS-TARGET-FRAMEWORK" for issue in result.issues)


def test_audit_orchestrator_readiness_rejects_entry_point_mismatch(tmp_path):
    manifest = _write_manifest(tmp_path / "manifest.json")

    def fake_run(command):
        if command[:3] == ["or", "processes", "get"]:
            return _result({
                "Key": command[3],
                "Name": "Invoice Processing",
                "ProcessKey": "InvoiceProcessing",
                "ProcessVersion": "1.2.4",
                "FolderPath": "Shared",
                "TargetFramework": "Windows",
                "EntryPointPath": "LegacyMain.xaml",
            })
        return _fake_run_ok([])(command)

    result = audit_orchestrator_readiness(manifest, run_uip=fake_run)

    assert not result.ok
    assert any(issue.code == "ORCH-PROCESS-ENTRY-POINT" for issue in result.issues)


def test_audit_orchestrator_readiness_rejects_non_rpa_process_type(tmp_path):
    manifest = _write_manifest(tmp_path / "manifest.json")

    def fake_run(command):
        if command[:3] == ["or", "processes", "get"]:
            return _result({
                "Key": command[3],
                "Name": "Invoice Processing",
                "ProcessKey": "InvoiceProcessing",
                "ProcessVersion": "1.2.4",
                "FolderPath": "Shared",
                "ProcessType": "Agent",
                "Enabled": True,
                "TargetFramework": "Windows",
                "EntryPointPath": "Main.xaml",
            })
        return _fake_run_ok([])(command)

    result = audit_orchestrator_readiness(manifest, run_uip=fake_run)

    assert not result.ok
    assert any(issue.code == "ORCH-PROCESS-TYPE" for issue in result.issues)


def test_audit_orchestrator_readiness_rejects_disabled_process(tmp_path):
    manifest = _write_manifest(tmp_path / "manifest.json")

    def fake_run(command):
        if command[:3] == ["or", "processes", "get"]:
            return _result({
                "Key": command[3],
                "Name": "Invoice Processing",
                "ProcessKey": "InvoiceProcessing",
                "ProcessVersion": "1.2.4",
                "FolderPath": "Shared",
                "ProcessType": "Process",
                "Enabled": False,
                "TargetFramework": "Windows",
                "EntryPointPath": "Main.xaml",
            })
        return _fake_run_ok([])(command)

    result = audit_orchestrator_readiness(manifest, run_uip=fake_run)

    assert not result.ok
    assert any(issue.code == "ORCH-PROCESS-DISABLED" for issue in result.issues)


def test_audit_orchestrator_readiness_rejects_missing_required_input_arguments(tmp_path):
    manifest = _write_manifest_with_required_inputs(tmp_path / "manifest.json")

    result = audit_orchestrator_readiness(manifest, run_uip=_fake_run_ok([]))

    assert not result.ok
    assert any(
        issue.code == "ORCH-INPUT-ARGUMENTS-MISSING"
        and "in_RequiredTicket" in issue.message
        for issue in result.issues
    )


def test_audit_orchestrator_readiness_rejects_missing_package_entry_point(tmp_path):
    manifest = _write_manifest(tmp_path / "manifest.json")

    def fake_run(command):
        if command[:3] == ["or", "packages", "entry-points"]:
            return _result([{"Path": "Other.xaml", "InputArguments": {}}])
        return _fake_run_ok([])(command)

    result = audit_orchestrator_readiness(manifest, run_uip=fake_run)

    assert not result.ok
    assert any(issue.code == "ORCH-PACKAGE-ENTRY-POINT-MISSING" for issue in result.issues)


def test_audit_orchestrator_readiness_rejects_required_input_missing_from_package_schema(tmp_path):
    manifest = _write_manifest_with_required_inputs(
        tmp_path / "manifest.json",
        input_arguments={"in_RequiredTicket": "S-123"},
    )

    def fake_run(command):
        if command[:3] == ["or", "packages", "entry-points"]:
            return _result([{"Path": "Main.xaml", "InputArguments": {"other": "String"}}])
        return _fake_run_ok([])(command)

    result = audit_orchestrator_readiness(manifest, run_uip=fake_run)

    assert not result.ok
    assert any(issue.code == "ORCH-PACKAGE-ENTRY-INPUT-MISSING" for issue in result.issues)


def test_audit_orchestrator_readiness_accepts_supplied_required_input_arguments(tmp_path):
    manifest = _write_manifest_with_required_inputs(
        tmp_path / "manifest.json",
        input_arguments={"in_RequiredTicket": "S-123"},
    )

    result = audit_orchestrator_readiness(manifest, run_uip=_fake_run_ok([]))

    assert result.ok


def test_audit_orchestrator_readiness_warns_when_required_inputs_come_from_file(tmp_path):
    manifest = _write_manifest_with_required_inputs(
        tmp_path / "manifest.json",
        input_file="input-arguments.json",
    )

    result = audit_orchestrator_readiness(manifest, run_uip=_fake_run_ok([]))

    assert result.ok
    assert any(issue.code == "ORCH-INPUT-ARGUMENTS-FILE" for issue in result.issues)


def test_audit_orchestrator_readiness_checks_declared_runtime_resources(tmp_path):
    manifest = _write_manifest_with_resources(tmp_path / "manifest.json")
    calls = []
    base = _fake_run_ok(calls)

    def fake_run(command):
        calls.append(command)
        if command[:3] == ["resource", "assets", "list"]:
            name = command[command.index("--name") + 1]
            if name == "ArquivoConfiguracao_Performer":
                return _result([{"Name": name, "ValueType": "Text"}])
            if name == "CredencialDB2":
                return _result([{"Name": name, "ValueType": "Credential"}])
            return _result([])
        if command[:3] == ["resource", "queues", "list"]:
            return _result([{"QueueDefinitionName": "Fila de Transacoes"}])
        if command[:3] == ["resource", "buckets", "list"]:
            return _result([{
                "Name": "anexos",
                "Identifier": "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
            }])
        if command[:3] == ["resource", "bucket-files", "list"]:
            return _result({"items": [{"fullPath": "/input/config.xlsx"}]})
        if command[:3] == ["or", "calendars", "list"]:
            return _result([{"Name": "Calendario DEV"}])
        calls.pop()
        return base(command)

    result = audit_orchestrator_readiness(manifest, run_uip=fake_run)

    assert result.ok
    assert [
        "resource", "assets", "list",
        "--name", "ArquivoConfiguracao_Performer",
        "--limit", "1000",
        "--output", "json",
        "--type", "Text",
        "--folder-path", "Shared",
    ] in calls
    assert [
        "resource", "queues", "list",
        "--name", "Fila de Transacoes",
        "--limit", "1000",
        "--output", "json",
        "--folder-path", "Shared",
    ] in calls
    assert [
        "resource", "bucket-files", "list",
        "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
        "--prefix", "input/config.xlsx",
        "--take-hint", "1000",
        "--output", "json",
        "--folder-path", "Shared",
    ] in calls
    assert [
        "or", "calendars", "list",
        "--limit", "1000",
        "--all-fields",
        "--output", "json",
    ] in calls


def test_audit_orchestrator_readiness_rejects_partial_required_resource_match(tmp_path):
    manifest = _write_manifest_with_resources(tmp_path / "manifest.json")
    calls = []
    base = _fake_run_ok(calls)

    def fake_run(command):
        calls.append(command)
        if command[:3] == ["resource", "assets", "list"]:
            name = command[command.index("--name") + 1]
            return _result([{"Name": name, "ValueType": "Text"}])
        if command[:3] == ["resource", "queues", "list"]:
            return _result([{"QueueDefinitionName": "Fila de Transacoes DEV"}])
        if command[:3] == ["resource", "buckets", "list"]:
            return _result([{
                "Name": "anexos",
                "Identifier": "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
            }])
        if command[:3] == ["resource", "bucket-files", "list"]:
            return _result({"items": [{"fullPath": "/input/config.xlsx"}]})
        if command[:3] == ["or", "calendars", "list"]:
            return _result([{"Name": "Calendario DEV"}])
        calls.pop()
        return base(command)

    result = audit_orchestrator_readiness(manifest, run_uip=fake_run)

    assert not result.ok
    assert any(issue.code == "ORCH-REQUIRED-QUEUE-MISSING" for issue in result.issues)


def test_audit_orchestrator_readiness_resolves_runtime_config_assets(tmp_path):
    config_path = tmp_path / "Config.xlsx"
    manifest = _write_manifest_with_config_hints(tmp_path / "manifest.json", config_path)
    calls = []
    base = _fake_run_ok(calls)

    def fake_run(command):
        calls.append(command)
        if command[:3] == ["resource", "assets", "list"]:
            name = command[command.index("--name") + 1]
            asset_types = {
                "ArquivoConfiguracao_Performer": "Text",
                "CredencialDB2_DEV": "Credential",
                "CredencialOffice_DEV": "Credential",
                "ApiEndpoint_DEV": "Text",
            }
            if name in asset_types:
                return _result([{
                    "Name": name,
                    "Key": f"{name}-key",
                    "ValueType": asset_types[name],
                }])
            return _result([])
        if command[:3] == ["resource", "assets", "get-asset-value"]:
            return _result({
                "Name": "ArquivoConfiguracao_Performer",
                "ValueType": "Text",
                "StringValue": str(config_path),
            })
        if command[:3] == ["resource", "queues", "list"]:
            return _result([{"QueueDefinitionName": "Fila DEV"}])
        if command[:3] == ["or", "calendars", "list"]:
            return _result([{"Name": "Calendario DEV"}])
        calls.pop()
        return base(command)

    result = audit_orchestrator_readiness(
        manifest,
        run_uip=fake_run,
        resolve_config_assets=True,
    )

    assert result.ok
    assert [
        "resource", "assets", "get-asset-value",
        "ArquivoConfiguracao_Performer-key",
        "--output", "json",
        "--folder-path", "Shared",
    ] in calls
    assert [
        "resource", "assets", "list",
        "--name", "CredencialDB2_DEV",
        "--limit", "1000",
        "--output", "json",
        "--type", "Credential",
        "--folder-path", "Shared",
    ] in calls
    assert [
        "resource", "queues", "list",
        "--name", "Fila DEV",
        "--limit", "1000",
        "--output", "json",
        "--folder-path", "Shared",
    ] in calls
    assert [
        "resource", "assets", "list",
        "--name", "CredencialOffice_DEV",
        "--limit", "1000",
        "--output", "json",
        "--type", "Credential",
        "--folder-path", "Shared",
    ] in calls
    assert [
        "or", "calendars", "list",
        "--limit", "1000",
        "--all-fields",
        "--output", "json",
    ] in calls


def test_audit_orchestrator_readiness_flags_runtime_and_resource_failures(tmp_path):
    manifest = _write_manifest(tmp_path / "manifest.json")

    def fake_run(command):
        if command[:2] == ["login", "status"]:
            return _result({"Status": "Logged in"})
        if command[:3] == ["login", "tenant", "list"]:
            return _result([{"TenantName": "RPA_Desenvolvimento"}])
        if command[:3] == ["login", "tenant", "set"]:
            return _result({"TenantName": command[3]})
        if command[:3] == ["or", "packages", "versions"]:
            return _result([{"Version": "1.2.4"}])
        if command[:3] == ["or", "processes", "get"]:
            return _result({
                "Name": "Invoice Processing",
                "ProcessKey": "InvoiceProcessing",
                "ProcessVersion": "1.2.3",
                "FolderPath": "Shared",
            })
        if command[:3] == ["or", "processes", "resources"]:
            return _result([
                {
                    "ResourceType": "Asset",
                    "ResourceName": "Credential",
                    "ValidationStatus": "Missing",
                    "ValidationMessage": "asset not found in folder",
                }
            ])
        if command[:3] == ["or", "folders", "runtimes"]:
            return _result([
                {"Type": "Unattended", "Total": 1, "Connected": 1, "Available": 0}
            ])
        if command[:3] == ["or", "machines", "list"]:
            return _result([
                {"Name": "dev-bot-01", "Key": "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa", "UnattendedSlots": 1}
            ])
        if command[:4] == ["or", "sessions", "unattended", "list"]:
            return _result([
                {"SessionId": 701, "RuntimeType": "Unattended", "State": "Available"}
            ])
        raise AssertionError(command)

    result = audit_orchestrator_readiness(manifest, run_uip=fake_run)

    assert not result.ok
    codes = {issue.code for issue in result.issues}
    assert "ORCH-PROCESS-VERSION" in codes
    assert "ORCH-RESOURCE-INVALID" in codes
    assert "ORCH-RUNTIME-AVAILABLE" in codes


def test_audit_orchestrator_readiness_flags_package_version_missing(tmp_path):
    manifest = _write_manifest(tmp_path / "manifest.json")

    def fake_run(command):
        if command[:2] == ["login", "status"]:
            return _result({"Status": "Logged in"})
        if command[:3] == ["login", "tenant", "list"]:
            return _result([{"TenantName": "RPA_Desenvolvimento"}])
        if command[:3] == ["login", "tenant", "set"]:
            return _result({"TenantName": command[3]})
        if command[:3] == ["or", "packages", "versions"]:
            return _result([{"Version": "1.2.3"}])
        if command[:3] == ["or", "processes", "get"]:
            return _result({
                "Name": "Invoice Processing",
                "ProcessKey": "InvoiceProcessing",
                "ProcessVersion": "1.2.4",
                "FolderPath": "Shared",
            })
        if command[:3] == ["or", "processes", "resources"]:
            return _result([])
        if command[:3] == ["or", "folders", "runtimes"]:
            return _result([
                {"Type": "Unattended", "Total": 1, "Connected": 1, "Available": 1}
            ])
        if command[:3] == ["or", "machines", "list"]:
            return _result([
                {"Name": "dev-bot-01", "Key": "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa", "UnattendedSlots": 1}
            ])
        if command[:4] == ["or", "sessions", "unattended", "list"]:
            return _result([
                {"SessionId": 701, "RuntimeType": "Unattended", "State": "Available"}
            ])
        raise AssertionError(command)

    result = audit_orchestrator_readiness(manifest, run_uip=fake_run)

    assert not result.ok
    assert any(issue.code == "ORCH-PACKAGE-VERSION-MISSING" for issue in result.issues)


def test_audit_orchestrator_readiness_flags_missing_machine_and_session(tmp_path):
    manifest = _write_manifest(tmp_path / "manifest.json")

    def fake_run(command):
        if command[:2] == ["login", "status"]:
            return _result({"Status": "Logged in"})
        if command[:3] == ["login", "tenant", "list"]:
            return _result([{"TenantName": "RPA_Desenvolvimento"}])
        if command[:3] == ["login", "tenant", "set"]:
            return _result({"TenantName": command[3]})
        if command[:3] == ["or", "packages", "versions"]:
            return _result([{"Version": "1.2.4"}])
        if command[:3] == ["or", "processes", "get"]:
            return _result({
                "Name": "Invoice Processing",
                "ProcessKey": "InvoiceProcessing",
                "ProcessVersion": "1.2.4",
                "FolderPath": "Shared",
            })
        if command[:3] == ["or", "processes", "resources"]:
            return _result([])
        if command[:3] == ["or", "folders", "runtimes"]:
            return _result([
                {"Type": "Unattended", "Total": 1, "Connected": 1, "Available": 1}
            ])
        if command[:3] == ["or", "machines", "list"]:
            return _result([])
        if command[:4] == ["or", "sessions", "unattended", "list"]:
            return _result([])
        raise AssertionError(command)

    result = audit_orchestrator_readiness(manifest, run_uip=fake_run)

    assert not result.ok
    codes = {issue.code for issue in result.issues}
    assert "ORCH-MACHINE-MISSING" in codes
    assert "ORCH-SESSION-MISSING" in codes


def test_manifest_requires_folder_binding(tmp_path):
    manifest = tmp_path / "manifest.json"
    manifest.write_text(
        json.dumps({
            "items": [{
                "packageId": "InvoiceProcessing",
                "packageVersion": "1.2.4",
                "processKey": "11111111-1111-1111-1111-111111111111",
            }],
        }),
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="folderPath"):
        load_readiness_manifest(manifest)


def test_format_readiness_result_declares_smoke_boundary(tmp_path):
    manifest = _write_manifest(tmp_path / "manifest.json")
    result = audit_orchestrator_readiness(manifest, run_uip=_fake_run_ok([]))

    text = format_readiness_result(result)

    assert "ORCHESTRATOR readiness: OK" in text
    assert "Runtime proof boundary" in text
    assert "jobs start" in text


def test_cli_audit_orchestrator_readiness_uses_official_runner(monkeypatch, tmp_path, capsys):
    manifest = _write_manifest(tmp_path / "manifest.json")
    monkeypatch.setattr("uip_engine.official_uip.run_official_uip", _fake_run_ok([]))
    args = cli.build_parser().parse_args([
        "audit-orchestrator-readiness",
        str(manifest),
    ])

    rc = cli._cmd_audit_orchestrator_readiness(args)

    out = capsys.readouterr().out
    assert rc == cli.EXIT_OK
    assert "ORCHESTRATOR readiness: OK" in out
