"""Read-only Orchestrator readiness checks for published UiPath processes.

This module is intentionally post-publish oriented. Source/package gates can
prove the package shape and dependency feed; they cannot prove that a concrete
Orchestrator process in a concrete folder is bound to that package version and
has runtime capacity. A small manifest supplies that operational binding.
"""
from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from .official_uip import OfficialUipResult
from .publish_dev import (
    DEV_TENANT,
    RunUip,
    _envelope_data,
    _extract_package_versions,
    _first_str,
    _records,
    ensure_login,
)


@dataclass(frozen=True)
class ReadinessIssue:
    code: str
    severity: str
    subject: str
    message: str

    @property
    def is_error(self) -> bool:
        return self.severity.upper() == "ERROR"

    def to_dict(self) -> dict[str, str]:
        return {
            "code": self.code,
            "severity": self.severity,
            "subject": self.subject,
            "message": self.message,
        }


@dataclass(frozen=True)
class ReadinessEntry:
    package_id: str
    package_version: str
    process_key: str
    folder_path: str | None = None
    folder_key: str | None = None
    runtime_type: str = "Unattended"
    process_name: str | None = None
    target_framework: str | None = None
    entry_point_path: str | None = None
    required_resources: tuple["RequiredResource", ...] = field(default_factory=tuple)
    runtime_resource_hints: tuple["RuntimeResourceHint", ...] = field(default_factory=tuple)
    required_input_arguments: tuple[str, ...] = field(default_factory=tuple)
    input_arguments: Any = None
    input_file: str | None = None

    @property
    def subject(self) -> str:
        return f"{self.package_id}@{self.package_version}"

    @property
    def folder_ref(self) -> str | None:
        return self.folder_key or self.folder_path


@dataclass(frozen=True)
class RequiredResource:
    resource_type: str
    name: str
    folder_path: str | None = None
    folder_key: str | None = None
    value_type: str | None = None
    bucket_name: str | None = None
    bucket_key: str | None = None
    path: str | None = None
    source: str | None = None

    @property
    def subject(self) -> str:
        detail = self.path or self.name
        return f"{self.resource_type}:{detail}"


@dataclass(frozen=True)
class RuntimeResourceHint:
    resource_type: str
    config_key: str
    source: str | None = None


@dataclass(frozen=True)
class ReadinessResult:
    tenant: str
    manifest_path: Path
    entries: tuple[ReadinessEntry, ...]
    issues: tuple[ReadinessIssue, ...] = field(default_factory=tuple)

    @property
    def ok(self) -> bool:
        return not any(issue.is_error for issue in self.issues)

    @property
    def error_count(self) -> int:
        return sum(1 for issue in self.issues if issue.is_error)

    @property
    def warn_count(self) -> int:
        return sum(1 for issue in self.issues if not issue.is_error)

    def to_dict(self) -> dict[str, Any]:
        return {
            "tenant": self.tenant,
            "manifestPath": str(self.manifest_path),
            "summary": {
                "entries": len(self.entries),
                "errors": self.error_count,
                "warnings": self.warn_count,
                "ok": self.ok,
            },
            "entries": [
                {
                    "packageId": entry.package_id,
                    "packageVersion": entry.package_version,
                    "processKey": entry.process_key,
                    "folderPath": entry.folder_path,
                    "folderKey": entry.folder_key,
                    "runtimeType": entry.runtime_type,
                    "processName": entry.process_name,
                    "targetFramework": entry.target_framework,
                    "entryPointPath": entry.entry_point_path,
                    "requiredResources": [
                        {
                            "type": resource.resource_type,
                            "name": resource.name,
                            "folderPath": resource.folder_path,
                            "folderKey": resource.folder_key,
                            "valueType": resource.value_type,
                            "bucketName": resource.bucket_name,
                            "bucketKey": resource.bucket_key,
                            "path": resource.path,
                            "source": resource.source,
                        }
                        for resource in entry.required_resources
                    ],
                    "runtimeResourceHints": [
                        {
                            "type": hint.resource_type,
                            "configKey": hint.config_key,
                            "source": hint.source,
                        }
                        for hint in entry.runtime_resource_hints
                    ],
                    "requiredInputArguments": list(entry.required_input_arguments),
                    "inputArguments": entry.input_arguments,
                    "inputFile": entry.input_file,
                }
                for entry in self.entries
            ],
            "issues": [issue.to_dict() for issue in self.issues],
        }


def load_readiness_manifest(path: Path) -> tuple[str | None, tuple[ReadinessEntry, ...]]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(payload, list):
        raw_items = payload
        tenant = None
    elif isinstance(payload, dict):
        raw_items = payload.get("items") or payload.get("processes") or payload.get("entries")
        tenant = _string_or_none(
            payload.get("tenant") or payload.get("devTenant") or payload.get("tenantName")
        )
    else:
        raise ValueError("manifest must be a JSON object or array")
    if not isinstance(raw_items, list) or not raw_items:
        raise ValueError("manifest must contain a non-empty items/processes array")

    entries: list[ReadinessEntry] = []
    for index, raw in enumerate(raw_items, start=1):
        if not isinstance(raw, dict):
            raise ValueError(f"manifest item #{index} must be an object")
        entry = _entry_from_raw(index, raw)
        entries.append(entry)
    return tenant, tuple(entries)


def audit_orchestrator_readiness(
    manifest_path: Path,
    *,
    run_uip: RunUip,
    tenant: str | None = None,
    check_packages: bool = True,
    check_resources: bool = True,
    check_runtimes: bool = True,
    check_machines: bool = True,
    check_sessions: bool = True,
    resolve_config_assets: bool = False,
    require_available_runtime: bool = True,
) -> ReadinessResult:
    manifest_tenant, entries = load_readiness_manifest(manifest_path)
    effective_tenant = tenant or manifest_tenant or DEV_TENANT
    ensure_login(run_uip, dev_tenant=effective_tenant)

    issues: list[ReadinessIssue] = []
    for entry in entries:
        issues.extend(_check_required_input_arguments(entry))
        if check_packages:
            issues.extend(_check_package_versions(entry, run_uip))
            issues.extend(_check_package_entry_points(entry, run_uip))
        issues.extend(_check_process_binding(entry, run_uip))
        if check_resources:
            issues.extend(_check_process_resources(entry, run_uip))
            issues.extend(_check_required_resources(entry, run_uip))
            if resolve_config_assets:
                issues.extend(_check_runtime_config_resources(entry, run_uip))
        if check_runtimes:
            issues.extend(_check_folder_runtimes(
                entry,
                run_uip,
                require_available_runtime=require_available_runtime,
            ))
        if check_machines:
            issues.extend(_check_folder_machines(entry, run_uip))
        if check_sessions:
            issues.extend(_check_unattended_sessions(entry, run_uip))

    return ReadinessResult(
        tenant=effective_tenant,
        manifest_path=manifest_path,
        entries=entries,
        issues=tuple(issues),
    )


def format_readiness_result(result: ReadinessResult) -> str:
    status = "OK" if result.ok else "FAIL"
    lines = [
        (
            f"ORCHESTRATOR readiness: {status}; "
            f"entries={len(result.entries)}; "
            f"errors={result.error_count}; warnings={result.warn_count}; "
            f"tenant={result.tenant}"
        )
    ]
    if not result.issues:
        lines.append(
            "  All manifest entries passed feed, process, Package Requirements, "
            "declared runtime resources, runtime, machine and session checks."
        )
    else:
        for issue in result.issues:
            lines.append(
                f"  - {issue.severity} {issue.code} {issue.subject}: {issue.message}"
            )
    lines.append("")
    lines.append("Runtime proof boundary:")
    lines.append(
        "  Read-only readiness cannot prove Robot package-cache extraction, "
        "interactive session health, target application login, or first-run "
        "workflow behavior."
    )
    lines.append(
        "  For 100% execution proof, run a controlled smoke job per process "
        "with `uip or jobs start <process-key> --wait-for-completion` in the "
        "same folder/account/machine context."
    )
    return "\n".join(lines)


def readiness_result_to_json(result: ReadinessResult) -> str:
    return json.dumps(result.to_dict(), indent=2, ensure_ascii=False)


def _entry_from_raw(index: int, raw: dict[str, Any]) -> ReadinessEntry:
    package_id = _required(raw, index, "packageId", "package_id", "package", "id")
    package_version = _required(raw, index, "packageVersion", "package_version", "version")
    process_key = _required(raw, index, "processKey", "process_key", "releaseKey", "release_key")
    folder_path = _optional(raw, "folderPath", "folder_path", "folder")
    folder_key = _optional(raw, "folderKey", "folder_key")
    runtime_type = _optional(raw, "runtimeType", "runtime_type") or "Unattended"
    process_name = _optional(raw, "processName", "process_name", "name")
    target_framework = _optional(
        raw,
        "targetFramework",
        "target_framework",
        "expectedTargetFramework",
        "expected_target_framework",
    )
    entry_point_path = _normalize_entry_point_path(_optional(
        raw,
        "entryPointPath",
        "entry_point_path",
        "expectedEntryPointPath",
        "expected_entry_point_path",
    ))
    required_resources = _required_resources_from_raw(index, raw)
    runtime_resource_hints = _runtime_resource_hints_from_raw(index, raw)
    required_input_arguments = _tuple_strings(_first_any(
        raw,
        "requiredInputArguments",
        "required_input_arguments",
        "requiredInputs",
    ))
    input_arguments = _first_any(raw, "inputArguments", "input_arguments")
    input_file = _optional(raw, "inputFile", "input_file")
    if not folder_path and not folder_key:
        raise ValueError(
            f"manifest item #{index} must define folderPath/folder_path or folderKey/folder_key"
        )
    return ReadinessEntry(
        package_id=package_id,
        package_version=package_version,
        process_key=process_key,
        folder_path=folder_path,
        folder_key=folder_key,
        runtime_type=runtime_type,
        process_name=process_name,
        target_framework=target_framework,
        entry_point_path=entry_point_path,
        required_resources=required_resources,
        runtime_resource_hints=runtime_resource_hints,
        required_input_arguments=required_input_arguments,
        input_arguments=input_arguments,
        input_file=input_file,
    )


def _runtime_resource_hints_from_raw(
    item_index: int,
    raw: dict[str, Any],
) -> tuple[RuntimeResourceHint, ...]:
    value = _first_any(raw, "runtimeResourceHints", "runtime_resource_hints")
    if value is None:
        return ()
    if not isinstance(value, list):
        raise ValueError(f"manifest item #{item_index} runtimeResourceHints must be a list")
    hints: list[RuntimeResourceHint] = []
    for hint_index, item in enumerate(value, start=1):
        if not isinstance(item, dict):
            raise ValueError(
                f"manifest item #{item_index} runtimeResourceHints #{hint_index} must be an object"
            )
        hints.append(RuntimeResourceHint(
            resource_type=_hint_resource_type(_required(
                item,
                hint_index,
                "type",
                "resourceType",
                "resource_type",
                "kind",
            )),
            config_key=_required(item, hint_index, "configKey", "config_key", "name"),
            source=_optional(item, "source", "from"),
        ))
    return tuple(hints)


def _required_resources_from_raw(
    item_index: int,
    raw: dict[str, Any],
) -> tuple[RequiredResource, ...]:
    value = _first_any(
        raw,
        "requiredResources",
        "required_resources",
        "runtimeResources",
        "runtime_resources",
    )
    if value is None:
        return ()
    if not isinstance(value, list):
        raise ValueError(f"manifest item #{item_index} requiredResources must be a list")
    resources: list[RequiredResource] = []
    for resource_index, item in enumerate(value, start=1):
        if not isinstance(item, dict):
            raise ValueError(
                f"manifest item #{item_index} requiredResources #{resource_index} must be an object"
            )
        raw_resource_type = _required(
            item,
            resource_index,
            "type",
            "resourceType",
            "resource_type",
            "kind",
        )
        resource_type = _resource_type(raw_resource_type)
        value_type = _optional(item, "valueType", "value_type", "assetType", "asset_type")
        if value_type is None:
            normalized_raw_type = _normalize_name(raw_resource_type)
            if normalized_raw_type in {"credential", "credentials"}:
                value_type = "Credential"
            elif normalized_raw_type in {"secret", "secrets"}:
                value_type = "Secret"
        name = _optional(item, "name", "resourceName", "resource_name")
        bucket_name = _optional(item, "bucketName", "bucket_name", "bucket")
        bucket_key = _optional(item, "bucketKey", "bucket_key")
        path = _optional(item, "path", "filePath", "file_path", "prefix")
        if resource_type == "bucketFile":
            if not path:
                raise ValueError(
                    f"manifest item #{item_index} requiredResources #{resource_index} "
                    "bucketFile requires path/filePath"
                )
            if not bucket_name and not bucket_key:
                raise ValueError(
                    f"manifest item #{item_index} requiredResources #{resource_index} "
                    "bucketFile requires bucketName/bucketKey"
                )
            name = name or path
        elif not name:
            raise ValueError(
                f"manifest item #{item_index} requiredResources #{resource_index} "
                "requires name/resourceName"
            )
        resources.append(RequiredResource(
            resource_type=resource_type,
            name=name,
            folder_path=_optional(item, "folderPath", "folder_path", "folder"),
            folder_key=_optional(item, "folderKey", "folder_key"),
            value_type=value_type,
            bucket_name=bucket_name,
            bucket_key=bucket_key,
            path=path,
            source=_optional(item, "source", "from"),
        ))
    return tuple(resources)


def _resource_type(raw: str) -> str:
    normalized = _normalize_name(raw)
    if normalized in {"asset", "assets"}:
        return "asset"
    if normalized in {"credential", "credentials"}:
        return "asset"
    if normalized in {"secret", "secrets"}:
        return "asset"
    if normalized in {"queue", "queues"}:
        return "queue"
    if normalized in {"bucket", "buckets", "storagebucket", "storagebuckets"}:
        return "bucket"
    if normalized in {"bucketfile", "bucketfiles", "storagefile", "storagefiles"}:
        return "bucketFile"
    if normalized in {"calendar", "calendars", "nonworkingdayscalendar"}:
        return "calendar"
    raise ValueError(f"unsupported required resource type: {raw}")


def _hint_resource_type(raw: str) -> str:
    normalized = _normalize_name(raw)
    if normalized in {"credential", "credentials"}:
        return "credential"
    if normalized in {"secret", "secrets"}:
        return "secret"
    return _resource_type(raw)


def _required(raw: dict[str, Any], index: int, *names: str) -> str:
    value = _optional(raw, *names)
    if value is None:
        joined = "/".join(names)
        raise ValueError(f"manifest item #{index} missing required field {joined}")
    return value


def _optional(raw: dict[str, Any], *names: str) -> str | None:
    lowered = {str(key).lower(): value for key, value in raw.items()}
    for name in names:
        value = raw.get(name)
        if value is None:
            value = lowered.get(name.lower())
        text = _string_or_none(value)
        if text is not None:
            return text
    return None


def _string_or_none(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _tuple_strings(value: Any) -> tuple[str, ...]:
    if value is None:
        return ()
    if isinstance(value, list):
        return tuple(str(item).strip() for item in value if str(item).strip())
    if isinstance(value, tuple):
        return tuple(str(item).strip() for item in value if str(item).strip())
    text = str(value).strip()
    return (text,) if text else ()


def _check_required_input_arguments(entry: ReadinessEntry) -> list[ReadinessIssue]:
    if not entry.required_input_arguments:
        return []
    if entry.input_file:
        return [_warn(
            "ORCH-INPUT-ARGUMENTS-FILE",
            entry,
            (
                "required input arguments are declared, but readiness cannot "
                f"inspect inputFile {entry.input_file}; smoke start will rely on that file"
            ),
        )]
    supplied = _input_argument_names(entry.input_arguments)
    missing = [name for name in entry.required_input_arguments if name not in supplied]
    if missing:
        return [_error(
            "ORCH-INPUT-ARGUMENTS-MISSING",
            entry,
            (
                "manifest is missing required inputArguments for job start: "
                + ", ".join(missing)
            ),
        )]
    return []


def _input_argument_names(value: Any) -> set[str]:
    if value is None:
        return set()
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return set()
        try:
            value = json.loads(stripped)
        except json.JSONDecodeError:
            return set()
    if isinstance(value, dict):
        return {str(key) for key in value}
    return set()


def _check_package_versions(entry: ReadinessEntry, run_uip: RunUip) -> list[ReadinessIssue]:
    try:
        data = _run_checked(
            run_uip,
            [
                "or", "packages", "versions", entry.package_id,
                "--limit", "1000",
                "--output", "json",
            ],
        )
        versions = _extract_package_versions(data)
    except Exception as exc:
        return [_error(
            "ORCH-PACKAGE-VERSIONS",
            entry,
            f"could not list package versions for {entry.package_id}: {exc}",
        )]
    if entry.package_version not in versions:
        available = ", ".join(versions) if versions else "(none returned)"
        return [_error(
            "ORCH-PACKAGE-VERSION-MISSING",
            entry,
            (
                f"{entry.package_id} version {entry.package_version} is not visible "
                f"in the tenant package feed. Available: {available}"
            ),
        )]
    return []


def _check_package_entry_points(entry: ReadinessEntry, run_uip: RunUip) -> list[ReadinessIssue]:
    try:
        data = _run_checked(
            run_uip,
            [
                "or", "packages", "entry-points",
                f"{entry.package_id}:{entry.package_version}",
                "--output", "json",
            ],
        )
    except Exception as exc:
        return [_error(
            "ORCH-PACKAGE-ENTRY-POINTS",
            entry,
            f"could not list package entry points for {entry.package_id}: {exc}",
        )]

    records = _records(data)
    if not records:
        return [_error(
            "ORCH-PACKAGE-ENTRY-POINTS",
            entry,
            "package entry-points lookup returned no records",
        )]

    issues: list[ReadinessIssue] = []
    matching_record: dict[str, Any] | None = None
    if entry.entry_point_path:
        expected = entry.entry_point_path
        for record in records:
            actual = _entry_point_path_from_record(record)
            if actual and _normalize_entry_point_path(actual) == expected:
                matching_record = record
                break
        if matching_record is None:
            available = ", ".join(
                sorted(
                    {
                        path for record in records
                        for path in [_entry_point_path_from_record(record)]
                        if path
                    }
                )
            ) or "(none returned)"
            issues.append(_error(
                "ORCH-PACKAGE-ENTRY-POINT-MISSING",
                entry,
                (
                    f"published package entry point {entry.entry_point_path} "
                    f"was not found. Available: {available}"
                ),
            ))
    elif len(records) == 1:
        matching_record = records[0]
    else:
        issues.append(_warn(
            "ORCH-PACKAGE-ENTRY-POINT-UNKNOWN",
            entry,
            "manifest does not define entryPointPath and package exposes multiple entry points",
        ))

    if entry.required_input_arguments and matching_record is not None:
        schema_names, schema_seen = _entry_point_input_argument_names(matching_record)
        if schema_seen:
            missing = [
                name for name in entry.required_input_arguments
                if name not in schema_names
            ]
            if missing:
                issues.append(_error(
                    "ORCH-PACKAGE-ENTRY-INPUT-MISSING",
                    entry,
                    (
                        "published package entry point schema does not expose "
                        f"required input argument(s): {', '.join(missing)}"
                    ),
                ))
        else:
            issues.append(_warn(
                "ORCH-PACKAGE-ENTRY-INPUT-UNKNOWN",
                entry,
                "package entry point response did not expose input argument schema",
            ))
    return issues


def _check_process_binding(entry: ReadinessEntry, run_uip: RunUip) -> list[ReadinessIssue]:
    try:
        data = _run_checked(
            run_uip,
            ["or", "processes", "get", entry.process_key, "--all-fields", "--output", "json"],
        )
    except Exception as exc:
        return [_error("ORCH-PROCESS-GET", entry, f"could not read process: {exc}")]

    record = _first_record(data)
    if record is None:
        return [_error("ORCH-PROCESS-GET", entry, "process lookup returned no record")]

    issues: list[ReadinessIssue] = []
    actual_version = _first_str(
        record,
        "ProcessVersion",
        "PackageVersion",
        "Version",
        "ReleaseVersion",
    )
    if actual_version != entry.package_version:
        issues.append(_error(
            "ORCH-PROCESS-VERSION",
            entry,
            (
                f"process {entry.process_key} is bound to version "
                f"{actual_version or '(missing)'}, expected {entry.package_version}"
            ),
        ))

    actual_package = _first_str(
        record,
        "PackageName",
        "PackageId",
        "PackageIdentifier",
        "PackageKey",
        "ProcessKey",
    )
    if actual_package and _normalize_name(actual_package) != _normalize_name(entry.package_id):
        issues.append(_error(
            "ORCH-PROCESS-PACKAGE",
            entry,
            (
                f"process {entry.process_key} points to package/process "
                f"{actual_package}, expected {entry.package_id}"
            ),
        ))
    elif not actual_package:
        issues.append(_warn(
            "ORCH-PROCESS-PACKAGE-UNKNOWN",
            entry,
            "process response did not expose a package id/name field to compare",
        ))

    actual_name = _first_str(record, "Name", "ReleaseName", "ProcessName")
    if entry.process_name and actual_name and actual_name != entry.process_name:
        issues.append(_error(
            "ORCH-PROCESS-NAME",
            entry,
            f"process name is {actual_name}, expected {entry.process_name}",
        ))

    actual_process_type = _first_str(record, "ProcessType", "ReleaseType", "Type")
    if actual_process_type and _normalize_name(actual_process_type) not in {
        "process",
        "rpa",
        "rpaprocess",
    }:
        issues.append(_error(
            "ORCH-PROCESS-TYPE",
            entry,
            f"process type is {actual_process_type}, expected Process/RPA",
        ))

    actual_status = _first_str(record, "Status", "ProcessStatus", "State", "Availability")
    if actual_status and _normalize_name(actual_status) in {
        "unavailable",
        "disabled",
        "inactive",
        "deleted",
        "missing",
        "invalid",
        "faulted",
    }:
        issues.append(_error(
            "ORCH-PROCESS-STATUS",
            entry,
            f"process status is {actual_status}, expected an available process",
        ))

    enabled = _explicit_bool(_first_any(record, "Enabled", "IsEnabled", "IsActive"))
    if enabled is False:
        issues.append(_error(
            "ORCH-PROCESS-DISABLED",
            entry,
            "process response marks the process as disabled/inactive",
        ))

    actual_target_framework = _first_str(
        record,
        "TargetFramework",
        "TargetFrameworkType",
        "ProjectTargetFramework",
        "ProcessTargetFramework",
    )
    if entry.target_framework:
        if actual_target_framework:
            if _normalize_process_target_framework(actual_target_framework) != (
                _normalize_process_target_framework(entry.target_framework)
            ):
                issues.append(_error(
                    "ORCH-PROCESS-TARGET-FRAMEWORK",
                    entry,
                    (
                        f"process TargetFramework is {actual_target_framework}, "
                        f"expected {entry.target_framework}"
                    ),
                ))
        else:
            issues.append(_warn(
                "ORCH-PROCESS-TARGET-FRAMEWORK-UNKNOWN",
                entry,
                "process response did not expose TargetFramework to compare",
            ))

    actual_entry_point = _first_str(
        record,
        "EntryPointPath",
        "EntryPoint",
        "Main",
        "MainPath",
    )
    if entry.entry_point_path:
        if actual_entry_point:
            if _normalize_entry_point_path(actual_entry_point) != entry.entry_point_path:
                issues.append(_error(
                    "ORCH-PROCESS-ENTRY-POINT",
                    entry,
                    (
                        f"process EntryPointPath is {actual_entry_point}, "
                        f"expected {entry.entry_point_path}"
                    ),
                ))
        else:
            issues.append(_warn(
                "ORCH-PROCESS-ENTRY-POINT-UNKNOWN",
                entry,
                "process response did not expose EntryPointPath to compare",
            ))

    actual_folder_path = _first_str(record, "FolderPath", "OrganizationUnitFullyQualifiedName")
    if entry.folder_path and actual_folder_path:
        if _normalize_folder(actual_folder_path) != _normalize_folder(entry.folder_path):
            issues.append(_error(
                "ORCH-PROCESS-FOLDER",
                entry,
                f"process folder path is {actual_folder_path}, expected {entry.folder_path}",
            ))
    actual_folder_key = _first_str(record, "FolderKey", "OrganizationUnitKey")
    if entry.folder_key and actual_folder_key:
        if actual_folder_key.lower() != entry.folder_key.lower():
            issues.append(_error(
                "ORCH-PROCESS-FOLDER",
                entry,
                f"process folder key is {actual_folder_key}, expected {entry.folder_key}",
            ))

    return issues


def _check_process_resources(entry: ReadinessEntry, run_uip: RunUip) -> list[ReadinessIssue]:
    try:
        data = _run_checked(
            run_uip,
            [
                "or", "processes", "resources", entry.process_key,
                "--all-fields",
                "--output", "json",
            ],
        )
    except Exception as exc:
        return [_error(
            "ORCH-PROCESS-RESOURCES",
            entry,
            f"could not inspect package requirements/resources: {exc}",
        )]

    issues: list[ReadinessIssue] = []
    for record in _records(data):
        status = _first_str(
            record,
            "ValidationStatus",
            "ValidationState",
            "Status",
            "ResourceStatus",
        )
        detail = _first_str(
            record,
            "ValidationError",
            "ValidationMessage",
            "Message",
            "ErrorMessage",
        )
        if _resource_status_ok(status, detail):
            continue
        resource_type = _first_str(record, "ResourceType", "Type", "Kind") or "Resource"
        resource_name = _first_str(record, "ResourceName", "Name", "DisplayName", "Resource") or "(unnamed)"
        issues.append(_error(
            "ORCH-RESOURCE-INVALID",
            entry,
            (
                f"{resource_type} {resource_name} validation status is "
                f"{status or '(missing)'}"
                + (f": {detail}" if detail else "")
            ),
        ))
    return issues


def _entry_point_path_from_record(record: dict[str, Any]) -> str | None:
    return _first_str(
        record,
        "Path",
        "EntryPointPath",
        "FilePath",
        "File",
        "Main",
    )


def _entry_point_input_argument_names(record: dict[str, Any]) -> tuple[set[str], bool]:
    raw = _first_any(
        record,
        "InputArguments",
        "InputArgumentSchema",
        "InputArgumentsSchema",
        "InputSchema",
        "Arguments",
    )
    if raw is None:
        return set(), False
    parsed = _parse_jsonish(raw)
    names = _argument_names(parsed)
    return names, True


def _parse_jsonish(value: Any) -> Any:
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return value
        try:
            return json.loads(text)
        except Exception:
            return value
    return value


def _argument_names(value: Any) -> set[str]:
    if isinstance(value, dict):
        names = {
            str(key).strip()
            for key in value
            if isinstance(key, str) and key.strip()
        }
        for nested_key in ("properties", "arguments", "input", "inputs"):
            nested = value.get(nested_key)
            if nested is not None and nested is not value:
                names.update(_argument_names(nested))
        name = _string_or_none(value.get("name"))
        if name:
            names.add(name)
        return names
    if isinstance(value, list):
        names: set[str] = set()
        for item in value:
            names.update(_argument_names(item))
        return names
    return set()


def _check_required_resources(entry: ReadinessEntry, run_uip: RunUip) -> list[ReadinessIssue]:
    issues: list[ReadinessIssue] = []
    for resource in entry.required_resources:
        if resource.resource_type == "asset":
            issues.extend(_check_required_asset(entry, resource, run_uip))
        elif resource.resource_type == "queue":
            issues.extend(_check_required_named_resource(
                entry,
                resource,
                run_uip,
                command_group="queues",
                name_fields=("QueueDefinitionName", "Name", "QueueName"),
                missing_code="ORCH-REQUIRED-QUEUE-MISSING",
            ))
        elif resource.resource_type == "bucket":
            issues.extend(_check_required_named_resource(
                entry,
                resource,
                run_uip,
                command_group="buckets",
                name_fields=("Name", "BucketName"),
                missing_code="ORCH-REQUIRED-BUCKET-MISSING",
            ))
        elif resource.resource_type == "bucketFile":
            issues.extend(_check_required_bucket_file(entry, resource, run_uip))
        elif resource.resource_type == "calendar":
            issues.extend(_check_required_calendar(entry, resource, run_uip))
    return issues


def _check_required_asset(
    entry: ReadinessEntry,
    resource: RequiredResource,
    run_uip: RunUip,
) -> list[ReadinessIssue]:
    command = [
        "resource", "assets", "list",
        "--name", resource.name,
        "--limit", "1000",
        "--output", "json",
    ]
    if resource.value_type:
        command.extend(["--type", resource.value_type])
    folder_issue = _append_resource_folder_arg(command, entry, resource)
    if folder_issue:
        return [folder_issue]
    try:
        data = _run_checked(run_uip, command)
    except Exception as exc:
        return [_required_resource_error(
            "ORCH-REQUIRED-ASSET-CHECK",
            entry,
            resource,
            f"could not list asset {resource.name}: {exc}",
        )]

    matches = _exact_name_matches(_records(data), resource.name, ("Name", "AssetName"))
    if not matches:
        folder = resource.folder_path or resource.folder_key or entry.folder_ref or "(folder missing)"
        return [_required_resource_error(
            "ORCH-REQUIRED-ASSET-MISSING",
            entry,
            resource,
            f"asset {resource.name} was not found in folder {folder}",
        )]
    if resource.value_type:
        expected = _normalize_name(resource.value_type)
        mismatches = [
            _first_str(record, "ValueType", "Type", "AssetType", "ValueTypeDisplayName")
            for record in matches
        ]
        if not any(value and _normalize_name(value) == expected for value in mismatches):
            actual = ", ".join(value for value in mismatches if value) or "(unknown)"
            return [_required_resource_error(
                "ORCH-REQUIRED-ASSET-TYPE",
                entry,
                resource,
                f"asset {resource.name} type is {actual}, expected {resource.value_type}",
            )]
    return []


def _check_required_named_resource(
    entry: ReadinessEntry,
    resource: RequiredResource,
    run_uip: RunUip,
    *,
    command_group: str,
    name_fields: tuple[str, ...],
    missing_code: str,
) -> list[ReadinessIssue]:
    command = [
        "resource", command_group, "list",
        "--name", resource.name,
        "--limit", "1000",
        "--output", "json",
    ]
    folder_issue = _append_resource_folder_arg(command, entry, resource)
    if folder_issue:
        return [folder_issue]
    try:
        data = _run_checked(run_uip, command)
    except Exception as exc:
        return [_required_resource_error(
            "ORCH-REQUIRED-RESOURCE-CHECK",
            entry,
            resource,
            f"could not list {resource.resource_type} {resource.name}: {exc}",
        )]
    if not _exact_name_matches(_records(data), resource.name, name_fields):
        folder = resource.folder_path or resource.folder_key or entry.folder_ref or "(folder missing)"
        return [_required_resource_error(
            missing_code,
            entry,
            resource,
            f"{resource.resource_type} {resource.name} was not found in folder {folder}",
        )]
    return []


def _check_required_bucket_file(
    entry: ReadinessEntry,
    resource: RequiredResource,
    run_uip: RunUip,
) -> list[ReadinessIssue]:
    bucket_key = resource.bucket_key
    bucket_subject = resource.bucket_name or resource.bucket_key or "(bucket missing)"
    if not bucket_key:
        bucket_resource = RequiredResource(
            resource_type="bucket",
            name=resource.bucket_name or "",
            folder_path=resource.folder_path,
            folder_key=resource.folder_key,
            source=resource.source,
        )
        command = [
            "resource", "buckets", "list",
            "--name", bucket_resource.name,
            "--limit", "1000",
            "--output", "json",
        ]
        folder_issue = _append_resource_folder_arg(command, entry, resource)
        if folder_issue:
            return [folder_issue]
        try:
            data = _run_checked(run_uip, command)
        except Exception as exc:
            return [_required_resource_error(
                "ORCH-REQUIRED-BUCKET-CHECK",
                entry,
                resource,
                f"could not list bucket {bucket_resource.name}: {exc}",
            )]
        matches = _exact_name_matches(_records(data), bucket_resource.name, ("Name", "BucketName"))
        if not matches:
            return [_required_resource_error(
                "ORCH-REQUIRED-BUCKET-MISSING",
                entry,
                resource,
                f"bucket {bucket_resource.name} was not found before file check",
            )]
        bucket_key = _first_str(matches[0], "Key", "Identifier", "BucketKey", "Id")
        if not bucket_key:
            return [_required_resource_error(
                "ORCH-REQUIRED-BUCKET-KEY",
                entry,
                resource,
                f"bucket {bucket_subject} was found but no key/identifier was returned",
            )]

    command = [
        "resource", "bucket-files", "list",
        bucket_key,
        "--prefix", resource.path or resource.name,
        "--take-hint", "1000",
        "--output", "json",
    ]
    folder_issue = _append_resource_folder_arg(command, entry, resource)
    if folder_issue:
        return [folder_issue]
    try:
        data = _run_checked(run_uip, command)
    except Exception as exc:
        return [_required_resource_error(
            "ORCH-REQUIRED-BUCKET-FILE-CHECK",
            entry,
            resource,
            f"could not list bucket file {resource.path}: {exc}",
        )]
    expected_path = (resource.path or resource.name).strip().lstrip("/")
    if not any(_bucket_file_path_matches(record, expected_path) for record in _records(data)):
        return [_required_resource_error(
            "ORCH-REQUIRED-BUCKET-FILE-MISSING",
            entry,
            resource,
            f"bucket {bucket_subject} does not expose file/prefix {resource.path or resource.name}",
        )]
    return []


def _check_required_calendar(
    entry: ReadinessEntry,
    resource: RequiredResource,
    run_uip: RunUip,
) -> list[ReadinessIssue]:
    try:
        data = _run_checked(
            run_uip,
            [
                "or", "calendars", "list",
                "--limit", "1000",
                "--all-fields",
                "--output", "json",
            ],
        )
    except Exception as exc:
        return [_required_resource_error(
            "ORCH-REQUIRED-CALENDAR-CHECK",
            entry,
            resource,
            f"could not list calendar {resource.name}: {exc}",
        )]
    if not _exact_name_matches(_records(data), resource.name, ("Name", "CalendarName")):
        return [_required_resource_error(
            "ORCH-REQUIRED-CALENDAR-MISSING",
            entry,
            resource,
            f"calendar {resource.name} was not found in tenant",
        )]
    return []


def _check_runtime_config_resources(entry: ReadinessEntry, run_uip: RunUip) -> list[ReadinessIssue]:
    if not entry.runtime_resource_hints:
        return []
    bootstrap_assets = [
        resource for resource in entry.required_resources
        if resource.resource_type == "asset"
        and (
            resource.name.startswith("ArquivoConfiguracao_")
            or resource.source == "source:ArquivoConfiguracao"
        )
    ]
    if not bootstrap_assets:
        return [_error(
            "ORCH-CONFIG-ASSET-MISSING",
            entry,
            (
                "runtimeResourceHints are present, but no ArquivoConfiguracao_* "
                "asset is declared in requiredResources"
            ),
        )]

    issues: list[ReadinessIssue] = []
    generated: list[RequiredResource] = []
    for asset in bootstrap_assets:
        config_value, value_issues = _read_text_asset_value(entry, asset, run_uip)
        issues.extend(value_issues)
        if not config_value:
            continue
        config_path = Path(os.path.expandvars(config_value)).expanduser()
        if not config_path.is_file():
            issues.append(_required_resource_error(
                "ORCH-CONFIG-FILE-MISSING",
                entry,
                asset,
                f"configuration file from asset value was not found: {config_path}",
            ))
            continue
        parsed, parse_issues = _load_runtime_config_file(config_path, entry, asset)
        issues.extend(parse_issues)
        if parsed is None:
            continue
        settings, asset_rows, credential_rows = parsed
        for hint in entry.runtime_resource_hints:
            value = settings.get(hint.config_key)
            if value is None or not str(value).strip():
                issues.append(_error(
                    "ORCH-CONFIG-KEY-MISSING",
                    entry,
                    (
                        f"configuration file {config_path} does not define a "
                        f"value for runtime resource key {hint.config_key}"
                    ),
                ))
                continue
            generated.append(_resource_from_config_hint(
                hint,
                str(value).strip(),
                source=f"config:{config_path}:{hint.config_key}",
            ))
        for asset_name, folder_path in asset_rows:
            generated.append(RequiredResource(
                resource_type="asset",
                name=asset_name,
                folder_path=folder_path or None,
                source=f"config:{config_path}:Assets",
            ))
        for credential_name, folder_path in credential_rows:
            generated.append(RequiredResource(
                resource_type="asset",
                name=credential_name,
                folder_path=folder_path or None,
                value_type="Credential",
                source=f"config:{config_path}:Credentials",
            ))

    generated = _dedupe_required_resources(generated)
    if generated:
        generated_entry = ReadinessEntry(
            package_id=entry.package_id,
            package_version=entry.package_version,
            process_key=entry.process_key,
            folder_path=entry.folder_path,
            folder_key=entry.folder_key,
            runtime_type=entry.runtime_type,
            process_name=entry.process_name,
            required_resources=tuple(generated),
            runtime_resource_hints=(),
        )
        issues.extend(_check_required_resources(generated_entry, run_uip))
    return issues


def _read_text_asset_value(
    entry: ReadinessEntry,
    resource: RequiredResource,
    run_uip: RunUip,
) -> tuple[str | None, list[ReadinessIssue]]:
    command = [
        "resource", "assets", "list",
        "--name", resource.name,
        "--limit", "1000",
        "--output", "json",
        "--type", resource.value_type or "Text",
    ]
    folder_issue = _append_resource_folder_arg(command, entry, resource)
    if folder_issue:
        return None, [folder_issue]
    try:
        data = _run_checked(run_uip, command)
    except Exception as exc:
        return None, [_required_resource_error(
            "ORCH-CONFIG-ASSET-CHECK",
            entry,
            resource,
            f"could not list bootstrap config asset {resource.name}: {exc}",
        )]
    matches = _exact_name_matches(_records(data), resource.name, ("Name", "AssetName"))
    if not matches:
        return None, [_required_resource_error(
            "ORCH-CONFIG-ASSET-NOT-FOUND",
            entry,
            resource,
            f"bootstrap config asset {resource.name} was not found",
        )]
    key = _first_str(matches[0], "Key", "AssetKey", "Identifier", "Id")
    if not key:
        return None, [_required_resource_error(
            "ORCH-CONFIG-ASSET-KEY",
            entry,
            resource,
            f"bootstrap config asset {resource.name} was found but no key was returned",
        )]
    value_command = [
        "resource", "assets", "get-asset-value",
        key,
        "--output", "json",
    ]
    folder_issue = _append_resource_folder_arg(value_command, entry, resource)
    if folder_issue:
        return None, [folder_issue]
    try:
        value_data = _run_checked(run_uip, value_command)
    except Exception as exc:
        return None, [_required_resource_error(
            "ORCH-CONFIG-ASSET-VALUE",
            entry,
            resource,
            f"could not read bootstrap config asset value {resource.name}: {exc}",
        )]
    record = _first_record(value_data)
    if record is None:
        return None, [_required_resource_error(
            "ORCH-CONFIG-ASSET-VALUE",
            entry,
            resource,
            f"bootstrap config asset value {resource.name} returned no record",
        )]
    value = _first_str(record, "StringValue", "Value", "TextValue", "stringValue", "value")
    if not value:
        return None, [_required_resource_error(
            "ORCH-CONFIG-ASSET-VALUE",
            entry,
            resource,
            f"bootstrap config asset {resource.name} has no text value",
        )]
    return value, []


def _load_runtime_config_file(
    config_path: Path,
    entry: ReadinessEntry,
    source_resource: RequiredResource,
) -> tuple[
    tuple[dict[str, str], list[tuple[str, str | None]], list[tuple[str, str | None]]] | None,
    list[ReadinessIssue],
]:
    if config_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return None, [_required_resource_error(
            "ORCH-CONFIG-FILE-TYPE",
            entry,
            source_resource,
            f"unsupported configuration file type for automated parsing: {config_path.suffix}",
        )]
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        return None, [_required_resource_error(
            "ORCH-CONFIG-PARSER",
            entry,
            source_resource,
            f"openpyxl is required to parse configuration workbooks: {exc}",
        )]
    try:
        workbook = load_workbook(config_path, read_only=True, data_only=True)
    except Exception as exc:
        return None, [_required_resource_error(
            "ORCH-CONFIG-PARSE",
            entry,
            source_resource,
            f"could not open configuration workbook {config_path}: {exc}",
        )]
    settings = _read_config_settings(workbook)
    asset_rows = _read_config_asset_rows(workbook)
    credential_rows = _read_config_credential_rows(workbook)
    return (settings, asset_rows, credential_rows), []


def _read_config_settings(workbook: Any) -> dict[str, str]:
    if "Settings" not in workbook.sheetnames:
        return {}
    sheet = workbook["Settings"]
    headers = _header_indexes(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ()))
    name_index = headers.get("name")
    value_index = headers.get("value")
    if name_index is None or value_index is None:
        return {}
    settings: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = _cell_text(row, name_index)
        value = _cell_text(row, value_index)
        if name and value is not None:
            settings[name] = value
    return settings


def _read_config_asset_rows(workbook: Any) -> list[tuple[str, str | None]]:
    if "Assets" not in workbook.sheetnames:
        return []
    sheet = workbook["Assets"]
    headers = _header_indexes(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ()))
    asset_index = headers.get("asset")
    folder_index = headers.get("orchestratorassetfolder")
    if asset_index is None:
        return []
    rows: list[tuple[str, str | None]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        asset = _cell_text(row, asset_index)
        if not asset:
            continue
        rows.append((asset, _cell_text(row, folder_index) if folder_index is not None else None))
    return rows


def _read_config_credential_rows(workbook: Any) -> list[tuple[str, str | None]]:
    if "Credentials" not in workbook.sheetnames:
        return []
    sheet = workbook["Credentials"]
    headers = _header_indexes(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ()))
    credential_index = (
        headers.get("credentialasset")
        if "credentialasset" in headers
        else headers.get("asset")
    )
    folder_index = headers.get("orchestratorassetfolder")
    if credential_index is None:
        return []
    rows: list[tuple[str, str | None]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        credential = _cell_text(row, credential_index)
        if not credential:
            continue
        rows.append((
            credential,
            _cell_text(row, folder_index) if folder_index is not None else None,
        ))
    return rows


def _header_indexes(row: tuple[Any, ...]) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for index, value in enumerate(row):
        text = str(value).strip().lower() if value is not None else ""
        if text:
            indexes[text.replace(" ", "")] = index
    return indexes


def _cell_text(row: tuple[Any, ...], index: int) -> str | None:
    if index >= len(row):
        return None
    value = row[index]
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _resource_from_config_hint(
    hint: RuntimeResourceHint,
    value: str,
    *,
    source: str,
) -> RequiredResource:
    value_type = None
    if hint.resource_type == "asset":
        value_type = "Text"
    if hint.resource_type == "credential":
        value_type = "Credential"
    if hint.resource_type == "secret":
        value_type = "Secret"
    return RequiredResource(
        resource_type=(
            "asset" if hint.resource_type in {"credential", "secret"} else hint.resource_type
        ),
        name=value,
        value_type=value_type,
        source=source,
    )


def _dedupe_required_resources(resources: list[RequiredResource]) -> list[RequiredResource]:
    seen: set[tuple[Any, ...]] = set()
    deduped: list[RequiredResource] = []
    for resource in resources:
        key = (
            resource.resource_type,
            resource.name.casefold(),
            (resource.folder_path or "").casefold(),
            (resource.folder_key or "").casefold(),
            (resource.value_type or "").casefold(),
            (resource.bucket_name or "").casefold(),
            (resource.bucket_key or "").casefold(),
            (resource.path or "").casefold(),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(resource)
    return deduped


def _check_folder_runtimes(
    entry: ReadinessEntry,
    run_uip: RunUip,
    *,
    require_available_runtime: bool,
) -> list[ReadinessIssue]:
    folder_ref = entry.folder_ref
    if not folder_ref:
        return [_error(
            "ORCH-FOLDER-RUNTIME",
            entry,
            "manifest entry does not define a folder path/key for runtime check",
        )]
    try:
        data = _run_checked(
            run_uip,
            ["or", "folders", "runtimes", folder_ref, "--output", "json"],
        )
    except Exception as exc:
        return [_error(
            "ORCH-FOLDER-RUNTIME",
            entry,
            f"could not list folder runtimes for {folder_ref}: {exc}",
        )]

    runtime = _find_runtime_record(data, entry.runtime_type)
    if runtime is None:
        return [_error(
            "ORCH-RUNTIME-MISSING",
            entry,
            f"folder {folder_ref} does not expose runtime type {entry.runtime_type}",
        )]

    connected = _int_field(runtime, "Connected", "ConnectedCount", "ConnectedRuntimes")
    available = _int_field(runtime, "Available", "AvailableCount", "AvailableRuntimes")
    total = _int_field(runtime, "Total", "TotalCount", "Runtimes")
    issues: list[ReadinessIssue] = []
    if total is not None and total <= 0:
        issues.append(_error(
            "ORCH-RUNTIME-TOTAL",
            entry,
            f"folder {folder_ref} has zero {entry.runtime_type} runtime slots",
        ))
    if connected is not None and connected <= 0:
        issues.append(_error(
            "ORCH-RUNTIME-CONNECTED",
            entry,
            f"folder {folder_ref} has zero connected {entry.runtime_type} runtimes",
        ))
    if require_available_runtime and available is not None and available <= 0:
        issues.append(_error(
            "ORCH-RUNTIME-AVAILABLE",
            entry,
            f"folder {folder_ref} has zero currently available {entry.runtime_type} runtimes",
        ))
    if total is None and connected is None and available is None:
        issues.append(_warn(
            "ORCH-RUNTIME-COUNTS-UNKNOWN",
            entry,
            f"runtime response for {entry.runtime_type} did not expose count fields",
        ))
    return issues


def _check_folder_machines(entry: ReadinessEntry, run_uip: RunUip) -> list[ReadinessIssue]:
    folder_ref = entry.folder_ref
    if not folder_ref:
        return [_error(
            "ORCH-MACHINE-FOLDER",
            entry,
            "manifest entry does not define a folder path/key for machine check",
        )]
    command = [
        "or", "machines", "list",
        "--all-fields",
        "--limit", "1000",
        "--output", "json",
    ]
    _append_folder_arg(command, entry)
    try:
        data = _run_checked(run_uip, command)
    except Exception as exc:
        return [_error(
            "ORCH-MACHINES",
            entry,
            f"could not list machines for folder {folder_ref}: {exc}",
        )]

    machines = _records(data)
    if not machines:
        return [_error(
            "ORCH-MACHINE-MISSING",
            entry,
            f"folder {folder_ref} has no assigned machines visible to the CLI session",
        )]

    slot_counts = [
        _runtime_slot_count(machine, entry.runtime_type)
        for machine in machines
    ]
    known_slot_counts = [count for count in slot_counts if count is not None]
    if known_slot_counts and max(known_slot_counts) <= 0:
        return [_error(
            "ORCH-MACHINE-RUNTIME-SLOTS",
            entry,
            (
                f"assigned machines for folder {folder_ref} expose zero "
                f"{entry.runtime_type} runtime slots"
            ),
        )]
    if not known_slot_counts:
        return [_warn(
            "ORCH-MACHINE-SLOTS-UNKNOWN",
            entry,
            (
                f"machines are assigned to folder {folder_ref}, but the response "
                f"did not expose {entry.runtime_type} slot fields"
            ),
        )]
    return []


def _check_unattended_sessions(entry: ReadinessEntry, run_uip: RunUip) -> list[ReadinessIssue]:
    folder_ref = entry.folder_ref
    if not folder_ref:
        return [_error(
            "ORCH-SESSION-FOLDER",
            entry,
            "manifest entry does not define a folder path/key for session check",
        )]
    command = [
        "or", "sessions", "unattended", "list",
        "--runtime-type", entry.runtime_type,
        "--limit", "1000",
        "--output", "json",
    ]
    _append_folder_arg(command, entry)
    try:
        data = _run_checked(run_uip, command)
    except Exception as exc:
        return [_error(
            "ORCH-SESSIONS",
            entry,
            f"could not list unattended sessions for folder {folder_ref}: {exc}",
        )]

    sessions = _records(data)
    if not sessions:
        return [_error(
            "ORCH-SESSION-MISSING",
            entry,
            (
                f"folder {folder_ref} has no unattended sessions for runtime "
                f"{entry.runtime_type}"
            ),
        )]

    states = [_session_state(session) for session in sessions]
    known_states = [state for state in states if state]
    if any(_session_usable(session, state) for session, state in zip(sessions, states)):
        return []
    if not known_states:
        return [_warn(
            "ORCH-SESSION-STATE-UNKNOWN",
            entry,
            (
                f"unattended sessions exist for folder {folder_ref}, but the "
                "response did not expose usable state fields"
            ),
        )]
    return [_error(
        "ORCH-SESSION-UNAVAILABLE",
        entry,
        (
            f"no usable unattended session for folder {folder_ref} and runtime "
            f"{entry.runtime_type}. Seen states: {', '.join(sorted(set(known_states)))}"
        ),
    )]


def _run_checked(run_uip: RunUip, command: list[str]) -> Any:
    result: OfficialUipResult = run_uip(command)
    return _envelope_data(result)


def _first_record(data: Any) -> dict[str, Any] | None:
    records = _records(data)
    return records[0] if records else None


def _resource_status_ok(status: str | None, detail: str | None) -> bool:
    if detail and detail.strip() and detail.strip() not in {"-", "None", "null"}:
        return False
    if status is None or not status.strip():
        return True
    normalized = status.strip().lower()
    return normalized in {
        "success",
        "succeeded",
        "valid",
        "available",
        "ok",
        "resolved",
        "present",
    }


def _find_runtime_record(data: Any, runtime_type: str) -> dict[str, Any] | None:
    expected = _normalize_name(runtime_type)
    for record in _records(data):
        actual = _first_str(record, "Type", "RuntimeType", "Name", "LicenseType")
        if actual and _normalize_name(actual) == expected:
            return record
    return None


def _int_field(record: dict[str, Any], *names: str) -> int | None:
    value = _first_any(record, *names)
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    match = re.search(r"-?\d+", str(value))
    return int(match.group(0)) if match else None


def _append_folder_arg(command: list[str], entry: ReadinessEntry) -> None:
    if entry.folder_key:
        command.extend(["--folder-key", entry.folder_key])
    elif entry.folder_path:
        command.extend(["--folder-path", entry.folder_path])


def _append_resource_folder_arg(
    command: list[str],
    entry: ReadinessEntry,
    resource: RequiredResource,
) -> ReadinessIssue | None:
    if resource.folder_key:
        command.extend(["--folder-key", resource.folder_key])
    elif resource.folder_path:
        command.extend(["--folder-path", resource.folder_path])
    elif entry.folder_key:
        command.extend(["--folder-key", entry.folder_key])
    elif entry.folder_path:
        command.extend(["--folder-path", entry.folder_path])
    else:
        return _required_resource_error(
            "ORCH-REQUIRED-RESOURCE-FOLDER",
            entry,
            resource,
            "required resource check needs folderPath/folderKey in the resource or entry",
        )
    return None


def _exact_name_matches(
    records: list[dict[str, Any]],
    expected_name: str,
    name_fields: tuple[str, ...],
) -> list[dict[str, Any]]:
    expected = _casefold_name(expected_name)
    return [
        record
        for record in records
        if any(
            (actual := _first_str(record, field)) is not None
            and _casefold_name(actual) == expected
            for field in name_fields
        )
    ]


def _bucket_file_path_matches(record: dict[str, Any], expected_path: str) -> bool:
    actual = _first_str(record, "FullPath", "Path", "Name", "ObjectName", "FilePath")
    if not actual:
        return False
    return actual.strip().lstrip("/").casefold() == expected_path.casefold()


def _casefold_name(value: str) -> str:
    return value.strip().casefold()


def _runtime_slot_count(record: dict[str, Any], runtime_type: str) -> int | None:
    normalized = _normalize_name(runtime_type)
    candidates = {
        "unattended": (
            "UnattendedSlots",
            "Unattended",
            "ProductionSlots",
            "Production",
            "UnattendedRuntimes",
        ),
        "nonproduction": (
            "NonProductionSlots",
            "NonProduction",
            "NonProductionRuntimes",
        ),
        "development": (
            "DevelopmentSlots",
            "Development",
            "DevelopmentRuntimes",
        ),
        "testautomation": (
            "TestAutomationSlots",
            "TestAutomation",
            "TestingSlots",
            "Testing",
        ),
        "headless": (
            "HeadlessSlots",
            "Headless",
            "HeadlessRuntimes",
        ),
    }
    names = candidates.get(normalized, (
        f"{runtime_type}Slots",
        runtime_type,
        f"{runtime_type}Runtimes",
    ))
    return _int_field(record, *names)


def _session_state(record: dict[str, Any]) -> str | None:
    state = _first_str(
        record,
        "State",
        "Status",
        "ConnectionState",
        "MachineState",
        "RuntimeState",
    )
    return state.strip() if state else None


def _session_usable(record: dict[str, Any], state: str | None) -> bool:
    if _bool_field(
        record,
        "IsUnresponsive",
        "Unresponsive",
        "IsDisconnected",
        "Disconnected",
        "IsInMaintenanceMode",
        "MaintenanceMode",
        "InMaintenanceMode",
    ):
        return False
    if state is None:
        return False
    normalized = _normalize_name(state)
    return normalized in {
        "available",
        "ready",
        "connected",
        "idle",
        "free",
    }


def _bool_field(record: dict[str, Any], *names: str) -> bool:
    value = _first_any(record, *names)
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "sim"}


def _explicit_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if value is None:
        return None
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "sim"}:
        return True
    if normalized in {"0", "false", "no", "nao", "não"}:
        return False
    return None


def _first_any(record: dict[str, Any], *names: str) -> Any:
    lowered = {str(key).lower(): value for key, value in record.items()}
    for name in names:
        if name in record:
            return record[name]
        value = lowered.get(name.lower())
        if value is not None:
            return value
    return None


def _normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def _normalize_folder(value: str) -> str:
    return re.sub(r"[\\/]+", "/", value.strip().strip("/")).lower()


def _normalize_entry_point_path(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = re.sub(r"[\\/]+", "/", value.strip().lstrip("/")).lower()
    return normalized or None


def _normalize_process_target_framework(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.strip().lower())


def _error(code: str, entry: ReadinessEntry, message: str) -> ReadinessIssue:
    return ReadinessIssue(code=code, severity="ERROR", subject=entry.subject, message=message)


def _warn(code: str, entry: ReadinessEntry, message: str) -> ReadinessIssue:
    return ReadinessIssue(code=code, severity="WARN", subject=entry.subject, message=message)


def _required_resource_error(
    code: str,
    entry: ReadinessEntry,
    resource: RequiredResource,
    message: str,
) -> ReadinessIssue:
    subject = f"{entry.subject} {resource.subject}"
    if resource.source:
        subject = f"{subject} [{resource.source}]"
    return ReadinessIssue(code=code, severity="ERROR", subject=subject, message=message)
